import json
import mimetypes
from io import BytesIO
from datetime import date, datetime, time
from decimal import Decimal, InvalidOperation
from pathlib import Path
from urllib.parse import quote

from django.conf import settings
from django.contrib.auth import authenticate, get_user_model, login, logout
from django.contrib.auth.models import Group
from django.core.files.base import ContentFile
from django.db import transaction
from django.db.models import Case, Count, F, IntegerField, Prefetch, Q, Sum, Value, When
from django.http import FileResponse, HttpResponse, HttpResponseNotAllowed, HttpResponseNotFound, JsonResponse
from django.shortcuts import get_object_or_404
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from .models import BusinessReview, ChangeRequest, Experiment, Invoice, LabDevice, LabOrder, LabStaffProfile, OrderDocument, OutsourceRequirement, ReportAudit, Sample, SamplePhoto, SchedulePlan, TestReport, TestStandard, WorkflowEvent
from .report_pdf import build_test_report_pdf


ROLE_SALES = '销售'
ROLE_BUSINESS = '商务'
ROLE_TECH = '技术'
ROLE_QUALITY = '质量部'
ROLE_SUZHOU_LAB = '苏州实验室'
ROLE_JIANGYIN_LAB = '江阴实验室'
ROLE_LAB_OPERATOR = '实验操作员'
ROLE_OUTSOURCE = '委外供应商'
ROLE_GENERAL_MANAGER = '总经理'
ROLE_ACCOUNTING = '会计'
ROLE_CHAIRMAN = '董事长'
VALID_ROLES = [
    ROLE_SALES,
    ROLE_BUSINESS,
    ROLE_TECH,
    ROLE_QUALITY,
    ROLE_SUZHOU_LAB,
    ROLE_JIANGYIN_LAB,
    ROLE_LAB_OPERATOR,
    ROLE_OUTSOURCE,
    ROLE_GENERAL_MANAGER,
    ROLE_ACCOUNTING,
    ROLE_CHAIRMAN,
]


def frontend(request):
    index_path = settings.BASE_DIR / 'static' / 'frontend' / 'index.html'
    if not index_path.exists():
        return HttpResponseNotFound(
            'Vue frontend has not been built. Run: cd frontend && pnpm build'
        )
    return HttpResponse(index_path.read_text(encoding='utf-8'))


def _is_chairman(user):
    return bool(
        user
        and user.is_authenticated
        and (user.is_superuser or user.groups.filter(name=ROLE_CHAIRMAN).exists())
    )


def _roles(user):
    if not user.is_authenticated:
        return []
    valid_role_set = set(VALID_ROLES)
    roles = [
        name
        for name in user.groups.values_list('name', flat=True)
        if name in valid_role_set
    ]
    if _is_chairman(user) and ROLE_CHAIRMAN not in roles:
        roles.insert(0, ROLE_CHAIRMAN)
    return roles


def _display_user(user):
    if not user:
        return ''
    return user.first_name or user.username


def _display_datetime(value):
    if not value:
        return ''
    if timezone.is_aware(value):
        value = timezone.localtime(value)
    return value.strftime('%Y-%m-%d %H:%M')


def _display_date(value):
    if not value:
        return ''
    if timezone.is_aware(value):
        value = timezone.localtime(value)
    return value.date().isoformat()


def _lab_profile(user):
    if not user or not user.is_authenticated:
        return None
    try:
        profile = user.lims_lab_profile
    except LabStaffProfile.DoesNotExist:
        return None
    return profile if profile.is_active else None


def _user_lab_type(user):
    profile = _lab_profile(user)
    if profile:
        return profile.lab_type
    roles = set(_roles(user))
    if ROLE_SUZHOU_LAB in roles:
        return LabDevice.LabType.SUZHOU
    if ROLE_JIANGYIN_LAB in roles:
        return LabDevice.LabType.JIANGYIN
    return None


def _is_lab_operator(user):
    profile = _lab_profile(user)
    return bool(ROLE_LAB_OPERATOR in _roles(user) and profile and profile.position == LabStaffProfile.Position.OPERATOR)


def _lab_schedule_query(lab_type):
    manager_role = ROLE_SUZHOU_LAB if lab_type == LabDevice.LabType.SUZHOU else ROLE_JIANGYIN_LAB
    return Q(test_type=lab_type) | Q(
        test_type=SchedulePlan.TestType.OUTSOURCE,
        lab_manager__groups__name=manager_role,
    )


def _can_operate_schedule(user, schedule):
    if _is_chairman(user):
        return True
    if _is_lab_operator(user):
        lab_type = _user_lab_type(user)
        if schedule.test_type == lab_type:
            return True
        manager_role = ROLE_SUZHOU_LAB if lab_type == LabDevice.LabType.SUZHOU else ROLE_JIANGYIN_LAB
        return bool(
            schedule.test_type == SchedulePlan.TestType.OUTSOURCE
            and schedule.lab_manager
            and schedule.lab_manager.groups.filter(name=manager_role).exists()
        )
    return schedule.lab_manager_id == user.id


def _can_view_finance(user):
    roles = set(_roles(user))
    return bool(_is_chairman(user) or ROLE_GENERAL_MANAGER in roles or ROLE_ACCOUNTING in roles)


def _user_payload(user):
    if not user.is_authenticated:
        return {'authenticated': False}
    roles = _roles(user)
    profile = _lab_profile(user)
    return {
        'authenticated': True,
        'id': user.id,
        'username': user.username,
        'email': user.email,
        'display_name': user.first_name or user.username,
        'roles': roles,
        'is_chairman': _is_chairman(user),
        'is_staff': user.is_staff,
        'is_superuser': user.is_superuser,
        'lab_type': profile.lab_type if profile else _user_lab_type(user),
        'lab_name': profile.get_lab_type_display() if profile else '',
        'lab_position': profile.get_position_display() if profile else '',
    }


def _orders_for_user(user):
    orders = LabOrder.objects.select_related(
        'sale_user', 'lead_lab_manager', 'outsource_requirement'
    ).prefetch_related('documents')
    if _is_chairman(user):
        return orders

    roles = set(_roles(user))
    query = Q()
    if ROLE_SALES in roles:
        query |= Q(sale_user=user)
    if ROLE_BUSINESS in roles or ROLE_TECH in roles:
        query |= Q(order_status__in=[LabOrder.Status.PENDING_REVIEW, LabOrder.Status.SCHEDULING])
    if ROLE_QUALITY in roles:
        query |= Q(order_status__in=[
            LabOrder.Status.SCHEDULING,
            LabOrder.Status.TESTING,
            LabOrder.Status.RESULT_PENDING,
            LabOrder.Status.TEST_FINISHED,
            LabOrder.Status.REPORT_REVIEW,
        ], workflow_version=LabOrder.WorkflowVersion.LEGACY_QUALITY)
    if ROLE_SUZHOU_LAB in roles or ROLE_JIANGYIN_LAB in roles:
        query |= Q(
            schedules__lab_manager=user,
            order_status__in=[
                LabOrder.Status.SCHEDULING,
                LabOrder.Status.TESTING,
                LabOrder.Status.RESULT_PENDING,
                LabOrder.Status.TEST_FINISHED,
                LabOrder.Status.REPORT_REVIEW,
            ],
        ) | Q(lead_lab_manager=user)
    if ROLE_LAB_OPERATOR in roles:
        lab_type = _user_lab_type(user)
        if lab_type:
            query |= Q(
                schedules__in=SchedulePlan.objects.filter(_lab_schedule_query(lab_type)),
                order_status__in=[
                    LabOrder.Status.SCHEDULING,
                    LabOrder.Status.TESTING,
                    LabOrder.Status.RESULT_PENDING,
                    LabOrder.Status.TEST_FINISHED,
                    LabOrder.Status.REPORT_REVIEW,
                ],
            )
    if ROLE_OUTSOURCE in roles:
        query |= Q(execution_mode__in=[LabOrder.ExecutionMode.OUTSOURCE, LabOrder.ExecutionMode.MIXED])
    if ROLE_GENERAL_MANAGER in roles:
        query |= Q()
        return orders
    if ROLE_ACCOUNTING in roles:
        business_pass_orders = BusinessReview.objects.filter(
            review_result=True,
            biz_review_user__isnull=False,
        ).values('order_id')
        tech_pass_orders = BusinessReview.objects.filter(
            review_result=True,
            tech_review_user__isnull=False,
        ).values('order_id')
        query |= (
            Q(id__in=business_pass_orders) & Q(id__in=tech_pass_orders)
        ) | Q(invoices__isnull=False)

    if not query:
        return orders.none()
    return orders.filter(query).distinct()


def _schedule_samples(schedule):
    prefetched = getattr(schedule, 'ordered_samples', None)
    if prefetched is not None:
        return prefetched
    return list(schedule.samples.select_related('quality_user', 'outbound_by').order_by('id'))


def _schedule_experiment(schedule):
    experiments = getattr(schedule, 'ordered_experiments', None)
    if experiments is not None:
        return experiments[0] if experiments else None
    return schedule.experiments.select_related('test_operator').order_by('-create_time').first()


def _sample_photo_payloads(schedule):
    photos = getattr(schedule, 'ordered_sample_photos', None)
    if photos is None:
        photos = schedule.sample_photos.all()
    return [
        {
            'id': photo.id,
            'name': photo.original_name,
            'size': photo.file_size,
            'url': f'/api/samples/photos/{photo.id}/',
        }
        for photo in photos
    ]


def _sample_lifecycle_payload(order):
    records = []
    for schedule in order.schedules.all():
        samples = _schedule_samples(schedule)
        photos = _sample_photo_payloads(schedule)
        if not samples:
            samples = [None]
        for sample in samples:
            actual_arrive_time = sample.actual_arrive_time if sample else schedule.sample_arrived_at
            records.append({
                'schedule_id': schedule.id,
                'sample_no': sample.sample_no if sample else '',
                'test_type': schedule.get_test_type_display(),
                'task_name': schedule.remark,
                'expected_arrive_time': _display_datetime(order.expect_sample_arrive),
                'actual_arrive_time': _display_datetime(actual_arrive_time),
                'outbound_time': _display_datetime(sample.outbound_time) if sample else '',
                'sample_status': sample.get_sample_status_display() if sample else ('样品已到' if schedule.sample_arrived else '样品未到'),
                'registered_by': _display_user(sample.quality_user) if sample else _display_user(schedule.sample_confirmed_by),
                'outbound_by': _display_user(sample.outbound_by) if sample else '',
                'photos': photos,
            })
    return records


def _experiment_lifecycle_payload(order):
    records = []
    for schedule in order.schedules.all():
        experiments = getattr(schedule, 'ordered_experiments', None)
        if experiments is None:
            experiments = schedule.experiments.select_related('test_operator').order_by('-create_time')
        for experiment in experiments:
            records.append({
                'schedule_id': schedule.id,
                'test_type': schedule.get_test_type_display(),
                'task_name': experiment.test_item_list or schedule.remark,
                'status': experiment.get_test_status_display(),
                'result_key': experiment.result_status,
                'result': experiment.get_result_status_display() if experiment.result_status else '',
                'started_at': _display_datetime(experiment.test_start_time),
                'ended_at': _display_datetime(experiment.test_end_time),
                'operator': _display_user(experiment.test_operator),
                'raw_data': experiment.test_raw_data,
                'conclusion': experiment.test_conclusion_temp,
            })
    return records


def _workflow_progress_payload(order):
    reviews = list(order.reviews.all())
    schedules = list(order.schedules.all())
    reports = list(order.reports.all())
    invoices = list(order.invoices.all())

    business_passes = [review for review in reviews if review.review_result and review.biz_review_user_id]
    technical_passes = [review for review in reviews if review.review_result and review.tech_review_user_id]
    latest_business = max(
        business_passes,
        key=lambda review: review.review_time or review.create_time,
        default=None,
    )
    latest_technical = max(
        technical_passes,
        key=lambda review: review.review_time or review.create_time,
        default=None,
    )
    latest_report = max(reports, key=lambda report: report.update_time or report.create_time, default=None)
    report_audits = list(latest_report.audits.all()) if latest_report else []
    sales_audits = [audit for audit in report_audits if audit.audit_level == ReportAudit.Level.SALES]
    gm_audits = [audit for audit in report_audits if audit.audit_level == ReportAudit.Level.GENERAL_MANAGER]
    latest_sales_audit = max(sales_audits, key=lambda audit: audit.audit_time, default=None)
    latest_gm_audit = max(gm_audits, key=lambda audit: audit.audit_time, default=None)
    final_invoice = max(
        (invoice for invoice in invoices if invoice.invoice_stage == Invoice.Stage.FINAL),
        key=lambda invoice: invoice.invoice_date or invoice.create_time,
        default=None,
    )
    preinvoices = [invoice for invoice in invoices if invoice.invoice_stage != Invoice.Stage.FINAL]
    is_v2 = order.workflow_version == LabOrder.WorkflowVersion.LAB_DIRECT

    schedule_count = len(schedules)
    planned_count = sum(bool(schedule.plan_start_time and schedule.plan_end_time) for schedule in schedules)
    arrived_count = 0
    ended_count = 0
    submitted_count = 0
    for schedule in schedules:
        samples = _schedule_samples(schedule)
        experiments = getattr(schedule, 'ordered_experiments', None)
        if experiments is None:
            experiments = list(schedule.experiments.all())
        arrived = schedule.sample_arrived or any(sample.actual_arrive_time for sample in samples)
        if arrived:
            arrived_count += 1
        if schedule.schedule_status in [SchedulePlan.Status.ENDED, SchedulePlan.Status.FINISHED] or any(
            experiment.test_status in [Experiment.Status.ENDED, Experiment.Status.FINISHED]
            for experiment in experiments
        ):
            ended_count += 1
        if schedule.schedule_status == SchedulePlan.Status.FINISHED or any(
            experiment.test_status == Experiment.Status.FINISHED for experiment in experiments
        ):
            submitted_count += 1

    advanced_to_scheduling = order.order_status in [
        LabOrder.Status.SCHEDULING,
        LabOrder.Status.TESTING,
        LabOrder.Status.RESULT_PENDING,
        LabOrder.Status.TEST_FINISHED,
        LabOrder.Status.REPORT_REVIEW,
        LabOrder.Status.INVOICED_CLOSED,
    ]
    advanced_to_testing = order.order_status in [
        LabOrder.Status.TESTING,
        LabOrder.Status.RESULT_PENDING,
        LabOrder.Status.TEST_FINISHED,
        LabOrder.Status.REPORT_REVIEW,
        LabOrder.Status.INVOICED_CLOSED,
    ]
    experiments_ended = order.order_status in [
        LabOrder.Status.RESULT_PENDING,
        LabOrder.Status.TEST_FINISHED,
        LabOrder.Status.REPORT_REVIEW,
        LabOrder.Status.INVOICED_CLOSED,
    ] or bool(schedule_count and ended_count == schedule_count)
    results_submitted = order.order_status in [
        LabOrder.Status.TEST_FINISHED,
        LabOrder.Status.REPORT_REVIEW,
        LabOrder.Status.INVOICED_CLOSED,
    ] or bool(schedule_count and submitted_count == schedule_count)

    steps = [
        {
            'key': 'sales_order', 'sequence': 1, 'phase': 'intake', 'phase_title': '业务准入',
            'title': '销售下单', 'owner': '销售', 'state': 'completed',
            'detail': f'{_display_user(order.sale_user) or "销售"}创建订单',
            'time': _display_datetime(order.create_time),
        },
        {
            'key': 'business_review', 'sequence': 2, 'phase': 'intake', 'phase_title': '业务准入',
            'title': '商务评审', 'owner': '商务', 'state': 'pending',
            'detail': _display_user(latest_business.biz_review_user) if latest_business else '核对报价、成本与交付条件',
            'time': _display_datetime(latest_business.review_time) if latest_business else '',
        },
        {
            'key': 'technical_review', 'sequence': 3, 'phase': 'intake', 'phase_title': '业务准入',
            'title': '技术评审', 'owner': '技术', 'state': 'pending',
            'detail': _display_user(latest_technical.tech_review_user) if latest_technical else '核对方法、标准与技术可行性',
            'time': _display_datetime(latest_technical.review_time) if latest_technical else '',
        },
        {
            'key': 'route_assignment', 'sequence': 4, 'phase': 'preparation', 'phase_title': '实施准备',
            'title': '路径与主责分配', 'owner': '技术' if is_v2 else '商务/质量部', 'state': 'pending',
            'detail': f'已分配{schedule_count}条执行路径' if schedule_count else '选择苏州、江阴、委外路径及主责负责人',
            'time': _display_datetime(min((schedule.create_time for schedule in schedules), default=None)),
        },
        {
            'key': 'scheduling', 'sequence': 5, 'phase': 'preparation', 'phase_title': '实施准备',
            'title': '排期排台', 'owner': '实验室' if is_v2 else '质量部', 'state': 'pending',
            'detail': f'{planned_count}/{schedule_count}条路径完成排期' if schedule_count else '等待执行路径分配',
            'time': _display_datetime(max((schedule.update_time for schedule in schedules), default=None)),
        },
        {
            'key': 'sales_confirmation', 'sequence': 6, 'phase': 'preparation', 'phase_title': '实施准备',
            'title': '销售确认需求', 'owner': '销售', 'state': 'pending',
            'detail': (
                '已确认样品与试验需求' if order.sales_confirmed_at
                else '流程已继续，历史记录未保存确认时间' if advanced_to_testing
                else '等待确认排期与需求是否变更'
            ),
            'time': _display_datetime(order.sales_confirmed_at),
        },
        {
            'key': 'sample_arrival', 'sequence': 7, 'phase': 'preparation', 'phase_title': '实施准备',
            'title': '样品到达确认', 'owner': '实验室' if is_v2 else '质量部', 'state': 'pending',
            'detail': f'{arrived_count}/{schedule_count}条路径已确认到样' if schedule_count else '等待排期后确认到样',
            'time': _display_datetime(max((schedule.sample_arrived_at for schedule in schedules if schedule.sample_arrived_at), default=None)),
        },
        {
            'key': 'experiment', 'sequence': 8, 'phase': 'execution', 'phase_title': '实验交付',
            'title': '实验执行', 'owner': '实验室/委外', 'state': 'pending',
            'detail': f'{ended_count}/{schedule_count}条路径实验已结束' if schedule_count else '等待样品与排期就绪',
            'time': _display_datetime(max((
                experiment.test_end_time
                for schedule in schedules
                for experiment in (getattr(schedule, 'ordered_experiments', None) or schedule.experiments.all())
                if experiment.test_end_time
            ), default=None)),
        },
        {
            'key': 'result_submission', 'sequence': 9, 'phase': 'execution', 'phase_title': '实验交付',
            'title': '提交实验结果', 'owner': '实验室/委外', 'state': 'pending',
            'detail': f'{submitted_count}/{schedule_count}条路径已提交结果' if schedule_count else '等待实验结束',
            'time': _display_datetime(max((
                schedule.update_time for schedule in schedules
                if schedule.schedule_status == SchedulePlan.Status.FINISHED
            ), default=None)),
        },
        {
            'key': 'report', 'sequence': 10, 'phase': 'execution', 'phase_title': '实验交付',
            'title': '出具检测报告', 'owner': '主责实验室' if is_v2 else '质量部', 'state': 'pending',
            'detail': f'{latest_report.report_no} · {latest_report.get_report_status_display()}' if latest_report else '等待全部结果提交',
            'time': _display_datetime(latest_report.update_time) if latest_report else '',
        },
        {
            'key': 'sales_audit', 'sequence': 11, 'phase': 'approval', 'phase_title': '审核结算',
            'title': '销售初审', 'owner': '销售', 'state': 'pending',
            'detail': latest_sales_audit.get_audit_result_display() if latest_sales_audit else '等待报告提交',
            'time': _display_datetime(latest_sales_audit.audit_time) if latest_sales_audit else '',
        },
        {
            'key': 'gm_audit', 'sequence': 12, 'phase': 'approval', 'phase_title': '审核结算',
            'title': '总经理终审', 'owner': '总经理', 'state': 'pending',
            'detail': latest_gm_audit.get_audit_result_display() if latest_gm_audit else '等待销售初审通过',
            'time': _display_datetime(latest_gm_audit.audit_time) if latest_gm_audit else '',
        },
        {
            'key': 'final_invoice', 'sequence': 13, 'phase': 'approval', 'phase_title': '审核结算',
            'title': '最终总开票', 'owner': '会计', 'state': 'pending',
            'detail': f'{final_invoice.invoice_no} · {final_invoice.invoice_amount:.2f}元' if final_invoice else '终审通过后开票并办结',
            'time': _display_datetime(final_invoice.invoice_date) if final_invoice else '',
        },
    ]
    step_map = {step['key']: step for step in steps}

    completed = {
        'sales_order': True,
        'business_review': bool(latest_business or advanced_to_scheduling),
        'technical_review': bool(latest_technical or advanced_to_scheduling),
        'route_assignment': bool(schedule_count or advanced_to_testing),
        'scheduling': bool((schedule_count and planned_count == schedule_count) or advanced_to_testing),
        'sales_confirmation': bool(order.sales_confirmed_at or advanced_to_testing),
        'sample_arrival': bool((schedule_count and arrived_count == schedule_count) or advanced_to_testing),
        'experiment': experiments_ended,
        'result_submission': results_submitted,
        'report': bool(latest_report and latest_report.report_status != TestReport.Status.DRAFT),
        'sales_audit': bool(latest_sales_audit and latest_sales_audit.audit_result == ReportAudit.Result.APPROVED),
        'gm_audit': bool(latest_gm_audit and latest_gm_audit.audit_result == ReportAudit.Result.APPROVED),
        'final_invoice': bool(final_invoice or order.order_status == LabOrder.Status.INVOICED_CLOSED),
    }
    for key, is_complete in completed.items():
        if is_complete:
            step_map[key]['state'] = 'completed'

    summary = order.get_order_status_display()
    current_keys = []
    if order.order_status == LabOrder.Status.CANCELLED:
        for step in steps[1:]:
            if step['state'] != 'completed':
                step['state'] = 'terminated'
        summary = '订单已退单，流程终止'
    elif order.order_status == LabOrder.Status.REVIEW_REJECTED:
        for key in ['business_review', 'technical_review']:
            if step_map[key]['state'] != 'completed':
                step_map[key]['state'] = 'rejected'
        summary = '评审驳回，等待销售修改或退单'
    elif latest_report and latest_report.report_status == TestReport.Status.REJECTED:
        if latest_gm_audit and latest_gm_audit.audit_result == ReportAudit.Result.REJECTED:
            step_map['gm_audit']['state'] = 'rejected'
        elif latest_sales_audit and latest_sales_audit.audit_result == ReportAudit.Result.REJECTED:
            step_map['sales_audit']['state'] = 'rejected'
        step_map['report']['state'] = 'current'
        current_keys = ['report']
        summary = '报告审核驳回，等待主责实验室重制'
    elif order.order_status == LabOrder.Status.PENDING_REVIEW:
        current_keys = [
            key for key in ['business_review', 'technical_review']
            if step_map[key]['state'] != 'completed'
        ] or ['route_assignment']
        summary = '商务与技术并行评审'
    elif order.order_status == LabOrder.Status.SCHEDULING:
        current_keys = [next(
            (key for key in ['route_assignment', 'scheduling', 'sales_confirmation', 'sample_arrival'] if step_map[key]['state'] != 'completed'),
            'experiment',
        )]
    elif order.order_status == LabOrder.Status.TESTING:
        current_keys = ['result_submission' if experiments_ended else 'experiment']
    elif order.order_status == LabOrder.Status.RESULT_PENDING:
        current_keys = ['result_submission']
        summary = '全部实验已结束，等待提交结果'
    elif order.order_status == LabOrder.Status.TEST_FINISHED:
        current_keys = ['report']
        summary = '实验结果已提交，等待出具报告'
    elif order.order_status == LabOrder.Status.REPORT_REVIEW:
        if not latest_report or latest_report.report_status == TestReport.Status.DRAFT:
            current_keys = ['report']
        elif latest_report.report_status == TestReport.Status.SALES_REVIEW:
            current_keys = ['sales_audit']
        elif latest_report.report_status == TestReport.Status.GM_REVIEW:
            current_keys = ['gm_audit']
        elif latest_report.report_status == TestReport.Status.APPROVED:
            current_keys = ['final_invoice']
    for key in current_keys:
        if step_map[key]['state'] != 'completed':
            step_map[key]['state'] = 'current'

    state_labels = {
        'completed': '已完成',
        'current': '当前处理',
        'pending': '待处理',
        'rejected': '已驳回',
        'terminated': '已终止',
    }
    for step in steps:
        step['state_label'] = state_labels[step['state']]

    current_titles = [step_map[key]['title'] for key in current_keys]
    return {
        'summary': summary,
        'current_step': '、'.join(current_titles) if current_titles else ('流程已完成' if completed['final_invoice'] else summary),
        'completed_steps': sum(step['state'] == 'completed' for step in steps),
        'total_steps': len(steps),
        'preinvoice_count': len(preinvoices),
        'preinvoice_total': str(sum((invoice.invoice_amount for invoice in preinvoices), Decimal('0.00'))),
        'steps': steps,
    }


def _outsource_requirement_payload(order):
    if not order.outsourced_execution:
        return None
    try:
        requirement = order.outsource_requirement
    except OutsourceRequirement.DoesNotExist:
        return None
    return {
        'outsource_company': requirement.outsource_company,
        'outsource_amount': str(requirement.outsource_amount),
        'entrust_order_no': requirement.entrust_order_no,
        'undertaking_amount': str(requirement.undertaking_amount),
        'experiment_start_time': _display_datetime(requirement.experiment_start_time),
        'experiment_end_time': _display_datetime(requirement.experiment_end_time),
        'created_by': _display_user(requirement.created_by),
        'created_at': _display_datetime(requirement.create_time),
    }


def _order_payload(order, include_sample_records=False):
    sample_arrival = order.expect_sample_arrive
    sample_arrival_value = _display_date(sample_arrival)
    delivery = order.expect_delivery_time
    if delivery:
        delivery_value = _display_date(delivery)
    else:
        delivery_value = ''

    execution_attributes = []
    if order.autonomous_execution:
        execution_attributes.append('自主')
    if order.outsourced_execution:
        execution_attributes.append('委外')

    payload = {
        'order_no': order.order_no,
        'is_outsource': order.outsourced_execution,
        'customer': order.customer_name,
        'contact': order.customer_contact,
        'phone': order.customer_phone,
        'project_name': order.project_name,
        'industry_category': order.industry_category,
        'industry_label': order.get_industry_category_display(),
        'test_demand': order.test_demand,
        'test_method': order.test_method,
        'test_standard': order.test_standard,
        'status': order.get_order_status_display(),
        'status_key': order.order_status,
        'execution_mode': order.get_execution_mode_display(),
        'execution_attributes': execution_attributes,
        'outsource_info': _outsource_requirement_payload(order),
        'workflow_version': order.workflow_version,
        'workflow_label': order.get_workflow_version_display(),
        'lead_lab_manager': _display_user(order.lead_lab_manager),
        'lead_lab_manager_username': order.lead_lab_manager.username if order.lead_lab_manager else '',
        'sales_confirmed': bool(order.sales_confirmed_at),
        'expected_sample_arrival': sample_arrival_value,
        'expected_delivery_date': delivery_value,
        'total_quote': str(order.total_quote),
        'is_urgent': order.is_urgent,
        'sales_owner': order.sale_user.first_name or order.sale_user.username
        if order.sale_user
        else '',
        'created_at': order.create_time.strftime('%Y-%m-%d %H:%M') if order.create_time else '',
        'remark': order.remark,
        'documents': [
            {
                'id': document.id,
                'type': document.document_type,
                'type_label': document.get_document_type_display(),
                'name': document.original_name,
                'size': document.file_size,
                'download_url': f'/api/orders/documents/{document.id}/download/',
            }
            for document in order.documents.all()
        ],
    }
    if include_sample_records:
        payload['sample_records'] = _sample_lifecycle_payload(order)
        payload['experiment_records'] = _experiment_lifecycle_payload(order)
        payload['workflow_progress'] = _workflow_progress_payload(order)
    return payload


def order_detail(request, order_no):
    if request.method != 'GET':
        return HttpResponseNotAllowed(['GET'])
    if not request.user.is_authenticated:
        return JsonResponse({'ok': False, 'error': '请先登录'}, status=401, json_dumps_params={'ensure_ascii': False})

    try:
        schedule_queryset = SchedulePlan.objects.select_related(
            'lab_manager', 'sample_confirmed_by'
        ).prefetch_related(
            Prefetch('sample_photos', queryset=SamplePhoto.objects.order_by('create_time', 'id'), to_attr='ordered_sample_photos'),
            Prefetch(
                'samples',
                queryset=Sample.objects.select_related('quality_user', 'outbound_by').order_by('id'),
                to_attr='ordered_samples',
            ),
            Prefetch(
                'experiments',
                queryset=Experiment.objects.select_related('test_operator').order_by('-create_time'),
                to_attr='ordered_experiments',
            ),
        )
        report_queryset = TestReport.objects.prefetch_related(
            Prefetch('audits', queryset=ReportAudit.objects.select_related('audit_user').order_by('audit_time', 'id')),
            'invoices',
        )
        order = _orders_for_user(request.user).prefetch_related(
            Prefetch('reviews', queryset=BusinessReview.objects.select_related('biz_review_user', 'tech_review_user')),
            Prefetch('schedules', queryset=schedule_queryset),
            Prefetch('reports', queryset=report_queryset),
            Prefetch('invoices', queryset=Invoice.objects.select_related('finance_user', 'report')),
        ).get(order_no=order_no)
    except LabOrder.DoesNotExist:
        return JsonResponse(
            {'ok': False, 'error': '订单不存在或当前岗位无权查看'},
            status=404,
            json_dumps_params={'ensure_ascii': False},
        )
    return JsonResponse({'ok': True, 'order': _order_payload(order, include_sample_records=True)}, json_dumps_params={'ensure_ascii': False})


def _report_payload(report):
    return {
        'id': report.id,
        'report_no': report.report_no,
        'order_no': report.order.order_no,
        'is_outsource': report.order.outsourced_execution,
        'customer': report.order.customer_name,
        'project_name': report.order.project_name,
        'status': report.get_report_status_display(),
        'status_key': report.report_status,
        'conclusion': report.final_conclusion,
        'remake_count': report.remake_count,
        'quality_user': report.create_quality_user.first_name or report.create_quality_user.username
        if report.create_quality_user
        else '',
        'report_type': report.report_type,
        'report_type_label': report.get_report_type_display(),
        'generated_at': report.generated_at.strftime('%Y-%m-%d %H:%M') if report.generated_at else '',
        'has_file': bool(report.report_file),
        'download_url': f'/api/reports/{report.id}/download/' if report.report_file else '',
    }


def _invoice_amounts(order):
    prefetched_invoices = getattr(order, '_prefetched_objects_cache', {}).get('invoices')
    if prefetched_invoices is not None:
        invoiced_total = sum((invoice.invoice_amount for invoice in prefetched_invoices), Decimal('0.00'))
    else:
        invoiced_total = order.invoices.aggregate(total=Sum('invoice_amount'))['total'] or Decimal('0.00')
    remaining_amount = max(Decimal('0.00'), order.total_quote - invoiced_total)
    return invoiced_total, remaining_amount


def _has_dual_review_pass(order):
    prefetched_reviews = getattr(order, '_prefetched_objects_cache', {}).get('reviews')
    if prefetched_reviews is not None:
        approved_reviews = [review for review in prefetched_reviews if review.review_result]
        return (
            any(review.biz_review_user_id for review in approved_reviews)
            and any(review.tech_review_user_id for review in approved_reviews)
        )
    reviews = order.reviews.filter(review_result=True)
    return reviews.filter(biz_review_user__isnull=False).exists() and reviews.filter(tech_review_user__isnull=False).exists()


def _preinvoice_stage(order):
    if not _has_dual_review_pass(order):
        return None
    if order.order_status in [LabOrder.Status.INVOICED_CLOSED, LabOrder.Status.CANCELLED]:
        return None
    prefetched_reports = getattr(order, '_prefetched_objects_cache', {}).get('reports')
    has_approved_report = (
        any(report.report_status == TestReport.Status.APPROVED for report in prefetched_reports)
        if prefetched_reports is not None
        else order.reports.filter(report_status=TestReport.Status.APPROVED).exists()
    )
    if has_approved_report:
        return None
    if order.order_status in [
        LabOrder.Status.RESULT_PENDING,
        LabOrder.Status.TEST_FINISHED,
        LabOrder.Status.REPORT_REVIEW,
    ]:
        return Invoice.Stage.PRE_EXPERIMENT
    if order.order_status in [LabOrder.Status.SCHEDULING, LabOrder.Status.TESTING]:
        return Invoice.Stage.PRE_REVIEW
    return None


def _experiment_finance_status(order):
    if order.order_status == LabOrder.Status.RESULT_PENDING:
        return '实验已结束，结果待提交'
    if order.order_status in [LabOrder.Status.TEST_FINISHED, LabOrder.Status.REPORT_REVIEW, LabOrder.Status.INVOICED_CLOSED]:
        return '实验结果已提交'
    prefetched_experiments = getattr(order, '_prefetched_objects_cache', {}).get('experiments')
    has_running_experiment = (
        any(experiment.test_status == Experiment.Status.RUNNING for experiment in prefetched_experiments)
        if prefetched_experiments is not None
        else order.experiments.filter(test_status=Experiment.Status.RUNNING).exists()
    )
    if has_running_experiment:
        return '实验进行中'
    return '实验未结束'


def _invoice_payload(invoice):
    order = invoice.order
    report = invoice.report
    invoiced_total, remaining_amount = _invoice_amounts(order)
    return {
        'invoice_no': invoice.invoice_no,
        'order_no': order.order_no,
        'is_outsource': order.outsourced_execution,
        'report_no': report.report_no if report else '',
        'customer': order.customer_name,
        'project_name': order.project_name,
        'invoice_amount': str(invoice.invoice_amount),
        'invoice_stage': invoice.invoice_stage,
        'invoice_stage_label': invoice.get_invoice_stage_display(),
        'order_total': str(order.total_quote),
        'invoiced_total': str(invoiced_total),
        'remaining_amount': str(remaining_amount),
        'invoice_type': invoice.invoice_type,
        'invoice_date': invoice.invoice_date.strftime('%Y-%m-%d') if invoice.invoice_date else '',
        'pay_status': invoice.get_pay_status_display(),
        'finish_status': invoice.get_order_finish_flag_display(),
        'finance_user': invoice.finance_user.first_name or invoice.finance_user.username
        if invoice.finance_user
        else '',
        'experiment_result_status': _experiment_finance_status(order),
    }


def _pending_final_invoice_payload(report):
    order = report.order
    invoiced_total, remaining_amount = _invoice_amounts(order)
    return {
        'invoice_no': '',
        'report_no': report.report_no,
        'order_no': order.order_no,
        'is_outsource': order.outsourced_execution,
        'customer': order.customer_name,
        'project_name': order.project_name,
        'invoice_amount': str(remaining_amount),
        'invoice_stage': Invoice.Stage.FINAL,
        'invoice_stage_label': Invoice.Stage.FINAL.label,
        'order_total': str(order.total_quote),
        'invoiced_total': str(invoiced_total),
        'remaining_amount': str(remaining_amount),
        'invoice_type': '待确认',
        'invoice_date': '',
        'pay_status': '待开票',
        'finish_status': order.get_order_status_display(),
        'finance_user': '',
        'experiment_result_status': _experiment_finance_status(order),
    }


def _pending_preinvoice_payload(order):
    stage = _preinvoice_stage(order)
    invoiced_total, remaining_amount = _invoice_amounts(order)
    return {
        'invoice_no': '',
        'report_no': '',
        'order_no': order.order_no,
        'is_outsource': order.outsourced_execution,
        'customer': order.customer_name,
        'project_name': order.project_name,
        'invoice_amount': '',
        'invoice_stage': stage,
        'invoice_stage_label': Invoice.Stage(stage).label if stage else '',
        'order_total': str(order.total_quote),
        'invoiced_total': str(invoiced_total),
        'remaining_amount': str(remaining_amount),
        'invoice_type': '待确认',
        'invoice_date': '',
        'pay_status': '待预开票',
        'finish_status': order.get_order_status_display(),
        'finance_user': '',
        'experiment_result_status': _experiment_finance_status(order),
    }


def _schedule_payload(schedule):
    order = schedule.order
    samples = _schedule_samples(schedule)
    sample = samples[0] if samples else None
    experiment = _schedule_experiment(schedule)
    return {
        'id': schedule.id,
        'order_no': order.order_no,
        'is_outsource': order.outsourced_execution,
        'customer': order.customer_name,
        'project_name': order.project_name,
        'status': order.get_order_status_display(),
        'status_key': order.order_status,
        'test_type': schedule.get_test_type_display(),
        'start_time': schedule.plan_start_time.strftime('%Y-%m-%d') if schedule.plan_start_time else '',
        'end_time': schedule.plan_end_time.strftime('%Y-%m-%d') if schedule.plan_end_time else '',
        'schedule_status': schedule.get_schedule_status_display(),
        'schedule_status_key': schedule.schedule_status,
        'lab_manager': _display_user(schedule.lab_manager),
        'device_id': schedule.device_id,
        'device_code': schedule.device.device_code if schedule.device else '',
        'device_name': schedule.device.device_name if schedule.device else '',
        'is_lead': bool(order.lead_lab_manager_id and order.lead_lab_manager_id == schedule.lab_manager_id),
        'sample_arrived': schedule.sample_arrived,
        'sample_arrival_status': '样品已到' if schedule.sample_arrived else '样品未到',
        'sample_arrived_at': _display_datetime(schedule.sample_arrived_at),
        'expected_sample_arrival': _display_datetime(order.expect_sample_arrive),
        'sample_outbound_at': _display_datetime(sample.outbound_time) if sample else '',
        'sample_status': sample.get_sample_status_display() if sample else ('样品已到' if schedule.sample_arrived else '样品未到'),
        'sample_photos': _sample_photo_payloads(schedule),
        'experiment_status': experiment.get_test_status_display() if experiment else '',
        'experiment_result_key': experiment.result_status if experiment else '',
        'experiment_result': experiment.get_result_status_display() if experiment and experiment.result_status else '',
        'experiment_conclusion': experiment.test_conclusion_temp if experiment else '',
        'experiment_raw_data': experiment.test_raw_data if experiment else '',
        'experiment_started_at': _display_datetime(experiment.test_start_time) if experiment else '',
        'experiment_ended_at': _display_datetime(experiment.test_end_time) if experiment else '',
        'experiment_operator': _display_user(experiment.test_operator) if experiment else '',
        'workflow_version': order.workflow_version,
        'remark': schedule.remark,
    }


def _sample_payload(sample):
    order = sample.order
    schedule = sample.schedule
    return {
        'sample_no': sample.sample_no,
        'order_no': order.order_no,
        'is_outsource': order.outsourced_execution,
        'customer': order.customer_name,
        'project_name': order.project_name,
        'sample_name': sample.sample_name,
        'sample_spec': sample.sample_spec,
        'sample_count': sample.sample_count,
        'storage_condition': sample.storage_condition,
        'expected_arrive_time': _display_datetime(order.expect_sample_arrive),
        'actual_arrive_time': _display_datetime(sample.actual_arrive_time),
        'outbound_time': _display_datetime(sample.outbound_time),
        'sample_status': sample.get_sample_status_display(),
        'test_type': schedule.get_test_type_display() if schedule else '',
        'quality_user': _display_user(sample.quality_user),
        'outbound_by': _display_user(sample.outbound_by),
        'photos': _sample_photo_payloads(schedule) if schedule else [],
    }


def _change_payload(change):
    order = change.order
    return {
        'order_no': order.order_no,
        'is_outsource': order.outsourced_execution,
        'customer': order.customer_name,
        'project_name': order.project_name,
        'scene': change.get_change_scene_display(),
        'status': change.get_change_status_display(),
        'content': change.change_content,
        'change_user': _display_user(change.change_user),
        'change_time': change.change_time.strftime('%Y-%m-%d %H:%M') if change.change_time else '',
    }


def _review_payload(review):
    order = review.order
    return {
        'order_no': order.order_no,
        'is_outsource': order.outsourced_execution,
        'customer': order.customer_name,
        'project_name': order.project_name,
        'biz_user': _display_user(review.biz_review_user),
        'tech_user': _display_user(review.tech_review_user),
        'result': '通过' if review.review_result else '驳回',
        'tech_feasible': '可行' if review.tech_feasible else '不可行',
        'reject_reason': review.reject_reason,
        'review_time': review.review_time.strftime('%Y-%m-%d %H:%M') if review.review_time else '',
    }


def _workflow_payload(event):
    order = event.order
    return {
        'order_no': order.order_no,
        'is_outsource': order.outsourced_execution,
        'customer': order.customer_name,
        'project_name': order.project_name,
        'actor': _display_user(event.actor),
        'event_type': event.get_event_type_display(),
        'from_status': event.from_status,
        'to_status': event.to_status,
        'note': event.note,
        'action_code': event.action_code,
        'schedule_id': event.schedule_id,
        'change_data': event.change_data,
        'change_summary': '；'.join(
            f"{item.get('label') or field}：{item.get('before', '')} → {item.get('after', '')}"
            for field, item in (event.change_data or {}).items()
            if isinstance(item, dict)
        ),
        'create_time': event.create_time.strftime('%Y-%m-%d %H:%M') if event.create_time else '',
    }


def _standard_payload(standard):
    return {
        'id': standard.id,
        'industry': standard.industry,
        'standard_code': standard.standard_code,
        'standard_name': standard.standard_name,
        'description': standard.description,
        'is_active': standard.is_active,
    }


def _role_user_options(role_name):
    return [
        {'id': user.id, 'username': user.username, 'name': _display_user(user)}
        for user in get_user_model().objects.filter(
            groups__name=role_name, is_active=True
        ).order_by('first_name', 'username')
    ]


def _limit_queryset(queryset, limit=50):
    try:
        limit_value = int(limit)
    except (TypeError, ValueError):
        limit_value = 50
    limit_value = max(1, min(limit_value, 200))
    return queryset[:limit_value]


def _device_booking_queryset(device, exclude_schedule_id=None):
    bookings = device.schedules.select_related('order', 'lab_manager', 'device').exclude(
        schedule_status=SchedulePlan.Status.FINISHED,
    ).exclude(
        order__order_status__in=[LabOrder.Status.INVOICED_CLOSED, LabOrder.Status.CANCELLED],
    )
    if exclude_schedule_id:
        bookings = bookings.exclude(id=exclude_schedule_id)
    return bookings


def _device_conflict(device, start_time, end_time, exclude_schedule_id=None):
    return _device_booking_queryset(device, exclude_schedule_id).filter(
        plan_start_time__lt=end_time,
        plan_end_time__gt=start_time,
    ).order_by('plan_start_time').first()


def _device_payload(device, start_time=None, end_time=None, exclude_schedule_id=None):
    now = timezone.now()
    bookings = _device_booking_queryset(device).order_by('plan_start_time', 'id')
    running = bookings.filter(schedule_status=SchedulePlan.Status.RUNNING).first()
    future = bookings.filter(plan_end_time__gte=now)
    if running:
        future = future.exclude(id=running.id)
    future = future[:8]
    conflict = None
    available = device.device_status == LabDevice.Status.NORMAL
    reason = ''
    if not available:
        reason = device.get_device_status_display()
    elif start_time and end_time:
        conflict = _device_conflict(device, start_time, end_time, exclude_schedule_id)
        if conflict:
            available = False
            reason = f'与 {conflict.order.order_no} 排期冲突'
    return {
        'id': device.id,
        'device_code': device.device_code,
        'name': device.device_name,
        'lab_type': device.lab_type,
        'lab_name': device.get_lab_type_display(),
        'model_spec': device.model_spec,
        'capability': device.capability,
        'status_key': device.device_status,
        'status': '实验中' if running else device.get_device_status_display(),
        'configured_status': device.get_device_status_display(),
        'remark': device.remark,
        'available': available,
        'unavailable_reason': reason,
        'order_no': running.order.order_no if running else '',
        'is_outsource': running.order.outsourced_execution if running else False,
        'project_name': running.order.project_name if running else '',
        'end_time': running.plan_end_time.strftime('%Y-%m-%d') if running and running.plan_end_time else '',
        'future_orders': [_schedule_payload(item) for item in future],
    }


def _lab_payload(test_type, name, related_orders=None, user=None):
    schedules = SchedulePlan.objects.select_related('order', 'device').prefetch_related(
        Prefetch('sample_photos', queryset=SamplePhoto.objects.order_by('create_time', 'id'), to_attr='ordered_sample_photos'),
        Prefetch(
            'samples',
            queryset=Sample.objects.select_related('quality_user', 'outbound_by').order_by('id'),
            to_attr='ordered_samples',
        ),
        Prefetch(
            'experiments',
            queryset=Experiment.objects.select_related('test_operator').order_by('-create_time'),
            to_attr='ordered_experiments',
        ),
    ).filter(test_type=test_type).order_by('plan_start_time')
    if related_orders is not None:
        schedules = schedules.filter(order__in=related_orders)
    if user is not None and not _is_chairman(user):
        roles = set(_roles(user))
        can_view_lab = ROLE_QUALITY in roles or ROLE_GENERAL_MANAGER in roles
        if ROLE_SUZHOU_LAB in roles and test_type == SchedulePlan.TestType.SUZHOU:
            can_view_lab = True
            schedules = schedules.filter(lab_manager=user)
        elif ROLE_JIANGYIN_LAB in roles and test_type == SchedulePlan.TestType.JIANGYIN:
            can_view_lab = True
            schedules = schedules.filter(lab_manager=user)
        elif ROLE_LAB_OPERATOR in roles and _user_lab_type(user) == test_type:
            can_view_lab = True
        if not can_view_lab:
            schedules = schedules.none()
    devices = LabDevice.objects.filter(lab_type=test_type).order_by('device_code')

    return {
        'name': name,
        'devices': [_device_payload(device) for device in devices],
        'orders': [_schedule_payload(item) for item in _limit_queryset(schedules, 80)],
    }


def _pending_reports_for_user(user, related_orders):
    reports = TestReport.objects.select_related('order', 'create_quality_user').filter(order__in=related_orders)
    if _is_chairman(user):
        return reports
    roles = set(_roles(user))
    query = Q()
    if ROLE_SALES in roles:
        query |= Q(order__sale_user=user)
    if ROLE_GENERAL_MANAGER in roles:
        return reports
    if ROLE_ACCOUNTING in roles:
        query |= Q(report_status=TestReport.Status.APPROVED)
    if ROLE_QUALITY in roles:
        query |= Q(order__workflow_version=LabOrder.WorkflowVersion.LEGACY_QUALITY)
    if ROLE_SUZHOU_LAB in roles or ROLE_JIANGYIN_LAB in roles:
        query |= Q(order__lead_lab_manager=user)
    if ROLE_LAB_OPERATOR in roles:
        lab_type = _user_lab_type(user)
        manager_role = ROLE_SUZHOU_LAB if lab_type == LabDevice.LabType.SUZHOU else ROLE_JIANGYIN_LAB
        query |= Q(order__lead_lab_manager__groups__name=manager_role)
    if not query:
        return reports.none()
    return reports.filter(query)


def _next_order_no():
    prefix = f'LIMS-{date.today().year}-'
    last_order = LabOrder.objects.filter(order_no__startswith=prefix).order_by('-order_no').first()
    if not last_order:
        return f'{prefix}0001'
    try:
        next_number = int(last_order.order_no.rsplit('-', 1)[1]) + 1
    except (IndexError, ValueError):
        next_number = LabOrder.objects.filter(order_no__startswith=prefix).count() + 1
    return f'{prefix}{next_number:04d}'


def _parse_datetime(value):
    if not value:
        return None
    try:
        parsed_date = date.fromisoformat(value)
        parsed = datetime.combine(parsed_date, time.min)
    except ValueError:
        try:
            parsed = datetime.fromisoformat(value)
        except ValueError:
            return None
    if timezone.is_naive(parsed):
        return timezone.make_aware(parsed)
    return parsed


def _parse_plan_range(start_value, end_value):
    start_time = _parse_datetime(start_value)
    end_time = _parse_datetime(end_value)
    if end_time and isinstance(end_value, str) and len(end_value.strip()) == 10:
        end_date = date.fromisoformat(end_value.strip())
        end_time = timezone.make_aware(datetime.combine(end_date, time.max))
    if not start_time or not end_time:
        return None, None, '计划开始和结束日期必填'
    if end_time <= start_time:
        return None, None, '计划结束日期必须晚于或等于开始日期'
    return start_time, end_time, ''


def _assign_schedule_device(schedule, payload, start_time, end_time):
    if schedule.test_type == SchedulePlan.TestType.OUTSOURCE:
        schedule.device = None
        return ''
    try:
        device_id = int(payload.get('device_id') or schedule.device_id or 0)
    except (TypeError, ValueError):
        device_id = 0
    if not device_id:
        return '内部实验室排期必须选择试验设备'
    device = LabDevice.objects.select_for_update().filter(id=device_id).first()
    if not device:
        return '选择的试验设备不存在'
    if device.lab_type != schedule.test_type:
        return '试验设备不属于当前实验室'
    if device.device_status != LabDevice.Status.NORMAL:
        return f'设备当前为“{device.get_device_status_display()}”，不可排期'
    conflict = _device_conflict(device, start_time, end_time, schedule.id)
    if conflict:
        conflict_end = conflict.plan_end_time.strftime('%Y-%m-%d') if conflict.plan_end_time else '待定'
        return f'设备与订单 {conflict.order.order_no} 的排期冲突（至 {conflict_end}）'
    schedule.device = device
    return ''


def _json_payload(request):
    if request.content_type and request.content_type.startswith('multipart/form-data'):
        return request.POST.dict(), None
    try:
        return json.loads(request.body.decode('utf-8') or '{}'), None
    except json.JSONDecodeError:
        return {}, JsonResponse({'ok': False, 'error': '请求格式错误'}, status=400, json_dumps_params={'ensure_ascii': False})


def _has_any_role(user, *roles):
    user_roles = set(_roles(user))
    return _is_chairman(user) or any(role in user_roles for role in roles)


def _require_auth(request):
    if not request.user.is_authenticated:
        return JsonResponse({'ok': False, 'error': '请先登录'}, status=401, json_dumps_params={'ensure_ascii': False})
    return None


def _require_role(user, *roles):
    if not _has_any_role(user, *roles):
        return JsonResponse({'ok': False, 'error': '当前岗位无权执行此操作'}, status=403, json_dumps_params={'ensure_ascii': False})
    return None


def _get_order(payload):
    order_no = (payload.get('order_no') or '').strip()
    if not order_no:
        return None, JsonResponse({'ok': False, 'error': '缺少订单号'}, status=400, json_dumps_params={'ensure_ascii': False})
    try:
        return LabOrder.objects.select_related('sale_user').prefetch_related('documents').get(order_no=order_no), None
    except LabOrder.DoesNotExist:
        return None, JsonResponse({'ok': False, 'error': '订单不存在'}, status=404, json_dumps_params={'ensure_ascii': False})


def _get_report(payload):
    report_no = (payload.get('report_no') or '').strip()
    if not report_no:
        return None, JsonResponse({'ok': False, 'error': '缺少报告号'}, status=400, json_dumps_params={'ensure_ascii': False})
    try:
        return TestReport.objects.select_related('order').get(report_no=report_no), None
    except TestReport.DoesNotExist:
        return None, JsonResponse({'ok': False, 'error': '报告不存在'}, status=404, json_dumps_params={'ensure_ascii': False})


def _status_response(message, order=None):
    payload = {'ok': True, 'message': message}
    if order:
        payload['order'] = _order_payload(order)
    return JsonResponse(payload, json_dumps_params={'ensure_ascii': False})


def _event(
    order,
    actor,
    note,
    from_status=None,
    to_status=None,
    event_type=WorkflowEvent.EventType.STATUS,
    action_code='',
    changes=None,
    schedule=None,
):
    WorkflowEvent.objects.create(
        order=order,
        actor=actor,
        event_type=event_type,
        from_status=str(from_status or ''),
        to_status=str(to_status or order.order_status or ''),
        note=note,
        action_code=action_code,
        change_data=changes or {},
        schedule=schedule,
    )


def _audit_change(label, before, after):
    def normalize(value):
        if value is None:
            return ''
        if isinstance(value, (datetime, date)):
            return value.isoformat()
        return str(value)
    return {'label': label, 'before': normalize(before), 'after': normalize(after)}


def _next_sample_no(order):
    return f'SMP-{order.order_no}-{order.samples.count() + 1:02d}'


def _next_report_no(order):
    return f'RPT-{order.order_no}-{order.reports.count() + 1:02d}'


def _next_invoice_no(order):
    return f'INV-{order.order_no}-{order.invoices.count() + 1:02d}'


def _first_user_in_group(role_name):
    return get_user_model().objects.filter(groups__name=role_name, is_active=True).first()


def _user_in_role(user_id, role_name):
    if not user_id:
        return None
    return get_user_model().objects.filter(
        id=user_id, groups__name=role_name, is_active=True
    ).first()


def _configure_v2_routes(order, actor, payload):
    routes = payload.get('execution_routes') or []
    if isinstance(routes, str):
        routes = [routes]
    routes = list(dict.fromkeys(routes))
    valid_routes = {'suzhou', 'jiangyin', 'outsource'}
    if not routes or not set(routes).issubset(valid_routes):
        return '技术评审必须选择至少一条有效执行路径'
    if ({'suzhou', 'jiangyin'} & set(routes)) and not order.autonomous_execution:
        return '销售订单未选择“自主”，不能分配内部实验室'
    if 'outsource' in routes and not order.outsourced_execution:
        return '销售订单未选择“委外”，不能分配委外路径'

    route_specs = []
    if 'suzhou' in routes:
        manager = _user_in_role(payload.get('suzhou_manager_id'), ROLE_SUZHOU_LAB)
        if not manager:
            return '请选择有效的苏州实验室负责人'
        route_specs.append((SchedulePlan.TestType.SUZHOU, manager, payload.get('suzhou_task') or order.test_demand))
    if 'jiangyin' in routes:
        manager = _user_in_role(payload.get('jiangyin_manager_id'), ROLE_JIANGYIN_LAB)
        if not manager:
            return '请选择有效的江阴实验室负责人'
        route_specs.append((SchedulePlan.TestType.JIANGYIN, manager, payload.get('jiangyin_task') or order.test_demand))
    if 'outsource' in routes:
        owner_id = payload.get('outsource_owner_id')
        manager = _user_in_role(owner_id, ROLE_SUZHOU_LAB) or _user_in_role(owner_id, ROLE_JIANGYIN_LAB)
        if not manager:
            return '请选择负责委外管理的内部实验室负责人'
        route_specs.append((SchedulePlan.TestType.OUTSOURCE, manager, payload.get('outsource_task') or order.test_demand))

    manager_ids = {manager.id for _, manager, _ in route_specs}
    try:
        lead_id = int(payload.get('lead_lab_manager_id') or 0)
    except (TypeError, ValueError):
        lead_id = 0
    if lead_id not in manager_ids:
        return '主责实验室负责人必须是本订单已分配的负责人之一'

    order.schedules.filter(samples__isnull=True, experiments__isnull=True).delete()
    for test_type, manager, task in route_specs:
        outsource_defaults = {}
        if test_type == SchedulePlan.TestType.OUTSOURCE:
            try:
                requirement = order.outsource_requirement
            except OutsourceRequirement.DoesNotExist:
                requirement = None
            if requirement:
                outsource_defaults = {
                    'outsource_factory': requirement.outsource_company,
                    'outsource_price': requirement.outsource_amount,
                    'plan_start_time': requirement.experiment_start_time,
                    'plan_end_time': requirement.experiment_end_time,
                }
        SchedulePlan.objects.create(
            order=order,
            test_type=test_type,
            lab_manager=manager,
            schedule_status=SchedulePlan.Status.NEW,
            quality_user=manager,
            assigned_by=actor,
            remark=task,
            **outsource_defaults,
        )
    order.lead_lab_manager_id = lead_id
    order.execution_mode = (
        route_specs[0][0] if len(route_specs) == 1 else LabOrder.ExecutionMode.MIXED
    )
    order.save(update_fields=['lead_lab_manager', 'execution_mode', 'update_time'])
    return ''


def _schedule_for_actor(order, payload, user):
    schedule_id = payload.get('schedule_id')
    schedules = order.schedules.all()
    if schedule_id:
        schedules = schedules.filter(id=schedule_id)
    if order.workflow_version == LabOrder.WorkflowVersion.LAB_DIRECT and not _is_chairman(user):
        if _is_lab_operator(user):
            lab_type = _user_lab_type(user)
            schedules = schedules.filter(_lab_schedule_query(lab_type)) if lab_type else schedules.none()
        else:
            schedules = schedules.filter(lab_manager=user)
    return schedules.order_by('id').first()


def current_user(request):
    return JsonResponse(_user_payload(request.user), json_dumps_params={'ensure_ascii': False})


@csrf_exempt
def lims_login(request):
    if request.method != 'POST':
        return HttpResponseNotAllowed(['POST'])
    try:
        payload = json.loads(request.body.decode('utf-8'))
    except json.JSONDecodeError:
        return JsonResponse({'ok': False, 'error': '请求格式错误'}, status=400, json_dumps_params={'ensure_ascii': False})

    username = payload.get('username', '').strip()
    password = payload.get('password', '')
    user = authenticate(request, username=username, password=password)
    if user is None:
        return JsonResponse({'ok': False, 'error': '用户名或密码错误'}, status=400, json_dumps_params={'ensure_ascii': False})
    if not user.is_active:
        return JsonResponse({'ok': False, 'error': '账号已停用'}, status=403, json_dumps_params={'ensure_ascii': False})

    login(request, user)
    return JsonResponse({'ok': True, 'user': _user_payload(user)}, json_dumps_params={'ensure_ascii': False})


@csrf_exempt
def lims_logout(request):
    if request.method != 'POST':
        return HttpResponseNotAllowed(['POST'])
    logout(request)
    return JsonResponse({'ok': True})


@csrf_exempt
@transaction.atomic
def add_employee(request):
    if request.method != 'POST':
        return HttpResponseNotAllowed(['POST'])
    if not _is_chairman(request.user):
        return JsonResponse({'ok': False, 'error': '仅董事长可以添加员工'}, status=403, json_dumps_params={'ensure_ascii': False})

    try:
        payload = json.loads(request.body.decode('utf-8'))
    except json.JSONDecodeError:
        return JsonResponse({'ok': False, 'error': '请求格式错误'}, status=400, json_dumps_params={'ensure_ascii': False})

    username = payload.get('username', '').strip()
    password = payload.get('password', '').strip()
    email = payload.get('email', '').strip()
    display_name = payload.get('display_name', '').strip()
    role = payload.get('role', '').strip()
    try:
        lab_type = int(payload.get('lab_type') or 0)
    except (TypeError, ValueError):
        lab_type = 0

    if not username or not password:
        return JsonResponse({'ok': False, 'error': '用户名和密码必填'}, status=400, json_dumps_params={'ensure_ascii': False})
    if role == ROLE_QUALITY:
        return JsonResponse({'ok': False, 'error': 'V2 工作流已取消质量部岗位，请选择实验室负责人'}, status=400, json_dumps_params={'ensure_ascii': False})
    if role == ROLE_LAB_OPERATOR and lab_type not in LabDevice.LabType.values:
        return JsonResponse({'ok': False, 'error': '实验操作员必须选择所属实验室'}, status=400, json_dumps_params={'ensure_ascii': False})

    user_model = get_user_model()
    if user_model.objects.filter(username=username).exists():
        return JsonResponse({'ok': False, 'error': '用户名已存在'}, status=400, json_dumps_params={'ensure_ascii': False})

    employee = user_model.objects.create_user(
        username=username,
        password=password,
        email=email,
        first_name=display_name,
        is_staff=role != ROLE_LAB_OPERATOR,
    )
    if role:
        group, _ = Group.objects.get_or_create(name=role)
        employee.groups.add(group)
    profile_position = None
    if role == ROLE_LAB_OPERATOR:
        profile_position = LabStaffProfile.Position.OPERATOR
    elif role == ROLE_SUZHOU_LAB:
        lab_type = LabDevice.LabType.SUZHOU
        profile_position = LabStaffProfile.Position.MANAGER
    elif role == ROLE_JIANGYIN_LAB:
        lab_type = LabDevice.LabType.JIANGYIN
        profile_position = LabStaffProfile.Position.MANAGER
    if profile_position:
        LabStaffProfile.objects.update_or_create(
            user=employee,
            defaults={'lab_type': lab_type, 'position': profile_position, 'is_active': True},
        )

    return JsonResponse(
        {
            'ok': True,
            'employee': {
                'id': employee.id,
                'username': employee.username,
                'display_name': employee.first_name or employee.username,
                'email': employee.email,
                'role': role,
                'lab_type': lab_type or None,
            },
        },
        json_dumps_params={'ensure_ascii': False},
    )


def _device_lab_types_for_user(user, include_read_only=False):
    if _is_chairman(user) or (include_read_only and ROLE_GENERAL_MANAGER in _roles(user)):
        return [LabDevice.LabType.SUZHOU, LabDevice.LabType.JIANGYIN]
    roles = set(_roles(user))
    lab_types = []
    if ROLE_SUZHOU_LAB in roles:
        lab_types.append(LabDevice.LabType.SUZHOU)
    if ROLE_JIANGYIN_LAB in roles:
        lab_types.append(LabDevice.LabType.JIANGYIN)
    return lab_types


@csrf_exempt
@transaction.atomic
def lab_devices(request):
    auth_error = _require_auth(request)
    if auth_error:
        return auth_error
    allowed_types = _device_lab_types_for_user(request.user, include_read_only=request.method == 'GET')
    if not allowed_types:
        return JsonResponse({'ok': False, 'error': '当前岗位无权访问设备管理'}, status=403, json_dumps_params={'ensure_ascii': False})
    if request.method == 'GET':
        devices = LabDevice.objects.filter(lab_type__in=allowed_types).order_by('lab_type', 'device_code')
        return JsonResponse(
            {'ok': True, 'devices': [_device_payload(device) for device in devices]},
            json_dumps_params={'ensure_ascii': False},
        )
    if request.method != 'POST':
        return HttpResponseNotAllowed(['GET', 'POST'])
    payload, parse_error = _json_payload(request)
    if parse_error:
        return parse_error
    try:
        lab_type = int(payload.get('lab_type') or (allowed_types[0] if len(allowed_types) == 1 else 0))
    except (TypeError, ValueError):
        lab_type = 0
    if lab_type not in allowed_types:
        return JsonResponse({'ok': False, 'error': '不能在其他实验室新增设备'}, status=403, json_dumps_params={'ensure_ascii': False})
    device_code = str(payload.get('device_code') or '').strip()
    device_name = str(payload.get('device_name') or '').strip()
    if not device_code or not device_name:
        return JsonResponse({'ok': False, 'error': '设备编号和设备名称必填'}, status=400, json_dumps_params={'ensure_ascii': False})
    if LabDevice.objects.filter(device_code=device_code).exists():
        return JsonResponse({'ok': False, 'error': '设备编号已存在'}, status=400, json_dumps_params={'ensure_ascii': False})
    device = LabDevice.objects.create(
        device_code=device_code,
        device_name=device_name,
        lab_type=lab_type,
        model_spec=str(payload.get('model_spec') or '').strip(),
        capability=str(payload.get('capability') or '').strip(),
        device_status=LabDevice.Status.NORMAL,
        remark=str(payload.get('remark') or '').strip(),
        created_by=request.user,
    )
    return JsonResponse(
        {'ok': True, 'message': '设备已新增', 'device': _device_payload(device)},
        status=201,
        json_dumps_params={'ensure_ascii': False},
    )


@csrf_exempt
@transaction.atomic
def lab_device_detail(request, device_id):
    auth_error = _require_auth(request)
    if auth_error:
        return auth_error
    allowed_types = _device_lab_types_for_user(request.user)
    device = get_object_or_404(LabDevice.objects.select_for_update(), id=device_id)
    if device.lab_type not in allowed_types:
        return JsonResponse({'ok': False, 'error': '无权管理该实验室设备'}, status=403, json_dumps_params={'ensure_ascii': False})
    if request.method == 'DELETE':
        if device.schedules.exists():
            return JsonResponse(
                {'ok': False, 'error': '设备已有排期或历史记录，不能删除；请改为“设备停用”'},
                status=400,
                json_dumps_params={'ensure_ascii': False},
            )
        device.delete()
        return JsonResponse({'ok': True, 'message': '设备已删除'}, json_dumps_params={'ensure_ascii': False})
    if request.method != 'PATCH':
        return HttpResponseNotAllowed(['PATCH', 'DELETE'])
    payload, parse_error = _json_payload(request)
    if parse_error:
        return parse_error
    try:
        status_value = int(payload.get('device_status') or device.device_status)
    except (TypeError, ValueError):
        status_value = 0
    if status_value not in LabDevice.Status.values:
        return JsonResponse({'ok': False, 'error': '设备状态无效'}, status=400, json_dumps_params={'ensure_ascii': False})
    if status_value != LabDevice.Status.NORMAL and device.schedules.filter(
        schedule_status=SchedulePlan.Status.RUNNING
    ).exists():
        return JsonResponse({'ok': False, 'error': '设备正在执行试验，不能直接维修或停用'}, status=400, json_dumps_params={'ensure_ascii': False})
    device.device_name = str(payload.get('device_name') or device.device_name).strip()
    device.model_spec = str(payload.get('model_spec') if payload.get('model_spec') is not None else device.model_spec).strip()
    device.capability = str(payload.get('capability') if payload.get('capability') is not None else device.capability).strip()
    device.remark = str(payload.get('remark') if payload.get('remark') is not None else device.remark).strip()
    device.device_status = status_value
    device.save()
    return JsonResponse(
        {'ok': True, 'message': '设备信息已更新', 'device': _device_payload(device)},
        json_dumps_params={'ensure_ascii': False},
    )


def lab_device_availability(request):
    auth_error = _require_auth(request)
    if auth_error:
        return auth_error
    role_error = _require_role(request.user, ROLE_SUZHOU_LAB, ROLE_JIANGYIN_LAB, ROLE_LAB_OPERATOR)
    if role_error:
        return role_error
    try:
        schedule_id = int(request.GET.get('schedule_id') or 0)
    except (TypeError, ValueError):
        schedule_id = 0
    schedule = SchedulePlan.objects.select_related('order').filter(id=schedule_id).first()
    if not schedule or schedule.test_type == SchedulePlan.TestType.OUTSOURCE:
        return JsonResponse({'ok': False, 'error': '内部实验室排期不存在'}, status=404, json_dumps_params={'ensure_ascii': False})
    if not _can_operate_schedule(request.user, schedule):
        return JsonResponse({'ok': False, 'error': '无权查询该排期设备'}, status=403, json_dumps_params={'ensure_ascii': False})
    start_time, end_time, range_error = _parse_plan_range(
        request.GET.get('start_date'), request.GET.get('end_date')
    )
    if range_error:
        return JsonResponse({'ok': False, 'error': range_error}, status=400, json_dumps_params={'ensure_ascii': False})
    devices = LabDevice.objects.filter(lab_type=schedule.test_type).order_by('device_code')
    return JsonResponse(
        {
            'ok': True,
            'schedule_id': schedule.id,
            'devices': [_device_payload(device, start_time, end_time, schedule.id) for device in devices],
        },
        json_dumps_params={'ensure_ascii': False},
    )


def _laboratory_schedule_queryset(request):
    try:
        lab_type = int(request.GET.get('lab_type') or _user_lab_type(request.user) or 0)
    except (TypeError, ValueError):
        lab_type = 0
    if lab_type not in LabDevice.LabType.values:
        return None, None, JsonResponse({'ok': False, 'error': '所属实验室无效'}, status=400, json_dumps_params={'ensure_ascii': False})
    allowed_lab_type = _user_lab_type(request.user)
    if not (_is_chairman(request.user) or ROLE_GENERAL_MANAGER in _roles(request.user)) and allowed_lab_type != lab_type:
        return None, None, JsonResponse({'ok': False, 'error': '无权查询其他实验室订单'}, status=403, json_dumps_params={'ensure_ascii': False})
    if not _has_any_role(request.user, ROLE_SUZHOU_LAB, ROLE_JIANGYIN_LAB, ROLE_LAB_OPERATOR, ROLE_GENERAL_MANAGER):
        return None, None, JsonResponse({'ok': False, 'error': '当前岗位无权查询实验室订单'}, status=403, json_dumps_params={'ensure_ascii': False})

    schedules = SchedulePlan.objects.select_related(
        'order', 'order__sale_user', 'lab_manager', 'device'
    ).prefetch_related(
        Prefetch('sample_photos', queryset=SamplePhoto.objects.order_by('create_time', 'id'), to_attr='ordered_sample_photos'),
        Prefetch(
            'samples',
            queryset=Sample.objects.select_related('quality_user', 'outbound_by').order_by('id'),
            to_attr='ordered_samples',
        ),
        Prefetch(
            'experiments',
            queryset=Experiment.objects.select_related('test_operator').order_by('-create_time'),
            to_attr='ordered_experiments',
        ),
    ).filter(_lab_schedule_query(lab_type)).distinct()
    keyword = (request.GET.get('keyword') or '').strip()
    if keyword:
        schedules = schedules.filter(
            Q(order__order_no__icontains=keyword)
            | Q(order__customer_name__icontains=keyword)
            | Q(order__project_name__icontains=keyword)
            | Q(remark__icontains=keyword)
            | Q(device__device_name__icontains=keyword)
        )
    if request.GET.get('order_status'):
        schedules = schedules.filter(order__order_status=request.GET['order_status'])
    if request.GET.get('schedule_status'):
        schedules = schedules.filter(schedule_status=request.GET['schedule_status'])
    if request.GET.get('device_id'):
        schedules = schedules.filter(device_id=request.GET['device_id'])
    start_time = _parse_datetime(request.GET.get('start_date'))
    end_time = _parse_datetime(request.GET.get('end_date'))
    if end_time:
        end_time += timezone.timedelta(days=1)
    if start_time:
        schedules = schedules.filter(plan_end_time__gte=start_time)
    if end_time:
        schedules = schedules.filter(plan_start_time__lt=end_time)
    selected_ids = [value for value in (request.GET.get('schedule_ids') or '').split(',') if value.isdigit()]
    if selected_ids:
        schedules = schedules.filter(id__in=selected_ids)
    schedules = schedules.annotate(
        operation_priority=Case(
            When(schedule_status=SchedulePlan.Status.CHANGE_PENDING, then=Value(0)),
            When(schedule_status=SchedulePlan.Status.NEW, sample_arrived=True, then=Value(1)),
            When(schedule_status=SchedulePlan.Status.RUNNING, then=Value(2)),
            When(schedule_status=SchedulePlan.Status.ENDED, then=Value(3)),
            When(schedule_status=SchedulePlan.Status.NEW, sample_arrived=False, then=Value(4)),
            When(schedule_status=SchedulePlan.Status.FINISHED, then=Value(5)),
            default=Value(6),
            output_field=IntegerField(),
        )
    )
    return schedules.order_by(
        'operation_priority',
        F('plan_start_time').asc(nulls_last=True),
        '-create_time',
        '-id',
    ), lab_type, None


def laboratory_orders(request):
    if request.method != 'GET':
        return HttpResponseNotAllowed(['GET'])
    auth_error = _require_auth(request)
    if auth_error:
        return auth_error
    schedules, lab_type, error = _laboratory_schedule_queryset(request)
    if error:
        return error
    try:
        page = max(1, int(request.GET.get('page') or 1))
        page_size = max(10, min(500, int(request.GET.get('page_size') or 100)))
    except (TypeError, ValueError):
        page, page_size = 1, 100
    total = schedules.count()
    start = (page - 1) * page_size
    items = schedules[start:start + page_size]
    return JsonResponse(
        {
            'ok': True,
            'lab_type': lab_type,
            'total': total,
            'page': page,
            'page_size': page_size,
            'items': [_schedule_payload(schedule) for schedule in items],
        },
        json_dumps_params={'ensure_ascii': False},
    )


def laboratory_orders_export(request):
    if request.method != 'GET':
        return HttpResponseNotAllowed(['GET'])
    auth_error = _require_auth(request)
    if auth_error:
        return auth_error
    schedules, lab_type, error = _laboratory_schedule_queryset(request)
    if error:
        return error

    workbook = Workbook()
    sheet = workbook.active
    lab_name = dict(LabDevice.LabType.choices)[lab_type]
    sheet.title = '实验室订单'
    headers = [
        '序号', '订单号', '客户名称', '项目名称', '试验任务', '执行路径', '设备编号', '设备名称',
        '计划开始', '计划结束', '预入库时间', '实际入库时间', '入库照片', '出库时间', '出库操作人',
        '样品状态', '排期状态', '订单状态', '负责人', '试验状态', '实验结果',
        '实际开始', '实际结束', '实验操作人', '实验结论', '销售', '报价', '预计交付',
    ]
    sheet.append(headers)
    header_fill = PatternFill('solid', fgColor='1F4E78')
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = Font(color='FFFFFF', bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    for index, schedule in enumerate(schedules[:5000], start=1):
        order = schedule.order
        experiment = schedule.ordered_experiments[0] if schedule.ordered_experiments else None
        samples = _schedule_samples(schedule)
        sample = samples[0] if samples else None
        photos = getattr(schedule, 'ordered_sample_photos', [])
        sheet.append([
            index,
            order.order_no,
            order.customer_name,
            order.project_name,
            schedule.remark,
            schedule.get_test_type_display(),
            schedule.device.device_code if schedule.device else '',
            schedule.device.device_name if schedule.device else '',
            schedule.plan_start_time.strftime('%Y-%m-%d') if schedule.plan_start_time else '',
            schedule.plan_end_time.strftime('%Y-%m-%d') if schedule.plan_end_time else '',
            _display_datetime(order.expect_sample_arrive),
            _display_datetime(sample.actual_arrive_time if sample else schedule.sample_arrived_at),
            '、'.join(photo.original_name for photo in photos),
            _display_datetime(sample.outbound_time) if sample else '',
            _display_user(sample.outbound_by) if sample else '',
            sample.get_sample_status_display() if sample else ('样品已到' if schedule.sample_arrived else '样品未到'),
            schedule.get_schedule_status_display(),
            order.get_order_status_display(),
            _display_user(schedule.lab_manager),
            experiment.get_test_status_display() if experiment else '',
            experiment.get_result_status_display() if experiment and experiment.result_status else '',
            _display_datetime(experiment.test_start_time) if experiment else '',
            _display_datetime(experiment.test_end_time) if experiment else '',
            _display_user(experiment.test_operator) if experiment else '',
            experiment.test_conclusion_temp if experiment else '',
            _display_user(order.sale_user),
            float(order.total_quote),
            order.expect_delivery_time.strftime('%Y-%m-%d') if order.expect_delivery_time else '',
        ])
    widths = [8, 20, 24, 30, 34, 18, 18, 22, 14, 14, 18, 18, 28, 18, 16, 14, 16, 16, 16, 16, 14, 18, 18, 16, 28, 14, 14, 14]
    for column, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(column)].width = width
    sheet.freeze_panes = 'A2'
    sheet.auto_filter.ref = sheet.dimensions
    output = BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = f'{lab_name}订单台账_{date.today().isoformat()}.xlsx'
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f"attachment; filename*=UTF-8''{quote(filename)}"
    return response


ORDER_DOCUMENT_EXTENSIONS = {'.doc', '.docx', '.pdf', '.jpg', '.jpeg', '.png'}
ORDER_DOCUMENT_MAX_SIZE = 20 * 1024 * 1024
ORDER_DOCUMENT_TOTAL_MAX_SIZE = 40 * 1024 * 1024


def _validate_order_documents(files, label, max_count):
    if len(files) > max_count:
        return f'{label}最多上传 {max_count} 个文件'
    for uploaded_file in files:
        extension = Path(uploaded_file.name).suffix.lower()
        if extension not in ORDER_DOCUMENT_EXTENSIONS:
            return f'{uploaded_file.name} 格式不支持，仅允许 Word、PDF、JPG、PNG'
        if uploaded_file.size > ORDER_DOCUMENT_MAX_SIZE:
            return f'{uploaded_file.name} 超过 20MB，请压缩后重新上传'
    return ''


def _request_order_payload(request):
    if request.content_type and request.content_type.startswith('multipart/form-data'):
        return request.POST, None
    try:
        return json.loads(request.body.decode('utf-8')), None
    except (json.JSONDecodeError, UnicodeDecodeError):
        return None, JsonResponse(
            {'ok': False, 'error': '请求格式错误'},
            status=400,
            json_dumps_params={'ensure_ascii': False},
        )


def _truthy(value):
    return str(value).strip().lower() in {'1', 'true', 'yes', 'on'}


@csrf_exempt
@transaction.atomic
def create_order(request):
    if request.method != 'POST':
        return HttpResponseNotAllowed(['POST'])
    if not request.user.is_authenticated:
        return JsonResponse({'ok': False, 'error': '请先登录'}, status=401, json_dumps_params={'ensure_ascii': False})

    roles = set(_roles(request.user))
    if not (_is_chairman(request.user) or ROLE_SALES in roles):
        return JsonResponse({'ok': False, 'error': '仅销售或董事长可以下单'}, status=403, json_dumps_params={'ensure_ascii': False})

    payload, parse_error = _request_order_payload(request)
    if parse_error:
        return parse_error

    customer_name = payload.get('customer_name', '').strip()
    project_name = payload.get('project_name', '').strip()
    test_demand = payload.get('test_requirements', '').strip()
    test_method = payload.get('test_method', '').strip()
    test_standard = payload.get('test_standard', '').strip()
    customer_contact = payload.get('contact_name', '').strip()
    customer_phone = payload.get('phone', '').strip()
    expect_sample_arrive = _parse_datetime(payload.get('expected_sample_arrival'))
    expect_delivery_time = _parse_datetime(payload.get('expected_delivery_date'))
    industry_category = (payload.get('industry_category') or '').strip()

    if hasattr(payload, 'getlist'):
        execution_attributes = payload.getlist('execution_attributes')
    else:
        execution_attributes = payload.get('execution_attributes') or ['autonomous']
    if isinstance(execution_attributes, str):
        execution_attributes = [execution_attributes]
    execution_attributes = set(execution_attributes)

    valid_attributes = {'autonomous', 'outsource'}
    if not execution_attributes or not execution_attributes.issubset(valid_attributes):
        return JsonResponse(
            {'ok': False, 'error': '订单执行属性至少选择“自主”或“委外”之一'},
            status=400,
            json_dumps_params={'ensure_ascii': False},
        )

    valid_industries = {choice.value for choice in LabOrder.IndustryCategory}
    if industry_category not in valid_industries:
        return JsonResponse(
            {'ok': False, 'error': '请选择行业属性：汽车、军工或其他'},
            status=400,
            json_dumps_params={'ensure_ascii': False},
        )

    contract_files = request.FILES.getlist('contract_files')
    outsource_contract_files = request.FILES.getlist('outsource_contract_files')
    attachment_files = request.FILES.getlist('attachment_files')
    file_error = _validate_order_documents(contract_files, '合同', 1)
    file_error = file_error or _validate_order_documents(outsource_contract_files, '委外合同', 1)
    file_error = file_error or _validate_order_documents(attachment_files, '附件', 10)
    all_order_files = contract_files + outsource_contract_files + attachment_files
    if not file_error and sum(item.size for item in all_order_files) > ORDER_DOCUMENT_TOTAL_MAX_SIZE:
        file_error = '合同、委外合同与附件总大小不能超过 40MB'
    if file_error:
        return JsonResponse(
            {'ok': False, 'error': file_error},
            status=400,
            json_dumps_params={'ensure_ascii': False},
        )

    if not customer_name or not project_name or not test_demand or not expect_sample_arrive:
        return JsonResponse({'ok': False, 'error': '客户名称、项目名称、试验需求、预计样品到达时间必填'}, status=400, json_dumps_params={'ensure_ascii': False})

    try:
        total_quote = Decimal(str(payload.get('quoted_amount') or '0'))
    except InvalidOperation:
        return JsonResponse({'ok': False, 'error': '报价金额格式错误'}, status=400, json_dumps_params={'ensure_ascii': False})

    outsource_requirement_data = None
    if 'outsource' in execution_attributes:
        outsource_company = (payload.get('outsource_company') or '').strip()
        entrust_order_no = (payload.get('entrust_order_no') or '').strip()
        experiment_start_time = _parse_datetime(payload.get('outsource_experiment_start_time'))
        experiment_end_time = _parse_datetime(payload.get('outsource_experiment_end_time'))
        if not all([
            outsource_contract_files,
            outsource_company,
            payload.get('outsource_amount'),
            entrust_order_no,
            payload.get('undertaking_amount'),
            experiment_start_time,
            experiment_end_time,
        ]):
            return JsonResponse(
                {'ok': False, 'error': '委外订单必须完整填写委外合同、委外金额、委外公司、委托单号、承接金额和实验起止时间'},
                status=400,
                json_dumps_params={'ensure_ascii': False},
            )
        try:
            outsource_amount = Decimal(str(payload.get('outsource_amount')))
            undertaking_amount = Decimal(str(payload.get('undertaking_amount')))
        except InvalidOperation:
            return JsonResponse({'ok': False, 'error': '委外金额或承接金额格式错误'}, status=400, json_dumps_params={'ensure_ascii': False})
        if (
            not outsource_amount.is_finite()
            or not undertaking_amount.is_finite()
            or outsource_amount <= 0
            or undertaking_amount <= 0
        ):
            return JsonResponse({'ok': False, 'error': '委外金额和承接金额必须大于 0'}, status=400, json_dumps_params={'ensure_ascii': False})
        if experiment_end_time <= experiment_start_time:
            return JsonResponse({'ok': False, 'error': '委外实验结束时间必须晚于开始时间'}, status=400, json_dumps_params={'ensure_ascii': False})
        outsource_requirement_data = {
            'outsource_company': outsource_company,
            'outsource_amount': outsource_amount,
            'entrust_order_no': entrust_order_no,
            'undertaking_amount': undertaking_amount,
            'experiment_start_time': experiment_start_time,
            'experiment_end_time': experiment_end_time,
        }
    elif outsource_contract_files:
        return JsonResponse({'ok': False, 'error': '只有选择“委外”属性的订单可以上传委外合同'}, status=400, json_dumps_params={'ensure_ascii': False})

    remark = '销售前台下单，等待商务技术评审。'
    if _truthy(payload.get('is_urgent')):
        remark = f'加急；{remark}'

    order = LabOrder.objects.create(
        order_no=_next_order_no(),
        customer_name=customer_name,
        customer_contact=customer_contact,
        customer_phone=customer_phone,
        project_name=project_name,
        industry_category=industry_category,
        test_demand=test_demand,
        test_method=test_method,
        test_standard=test_standard,
        sale_user=request.user,
        order_status=LabOrder.Status.PENDING_REVIEW,
        expect_sample_arrive=expect_sample_arrive,
        expect_delivery_time=expect_delivery_time,
        total_quote=total_quote,
        autonomous_execution='autonomous' in execution_attributes,
        outsourced_execution='outsource' in execution_attributes,
        create_by=request.user.username,
        update_by=request.user.username,
        remark=remark,
    )
    if outsource_requirement_data:
        OutsourceRequirement.objects.create(
            order=order,
            created_by=request.user,
            updated_by=request.user,
            **outsource_requirement_data,
        )
    for document_type, files in (
        (OrderDocument.DocumentType.CONTRACT, contract_files),
        (OrderDocument.DocumentType.OUTSOURCE_CONTRACT, outsource_contract_files),
        (OrderDocument.DocumentType.ATTACHMENT, attachment_files),
    ):
        for uploaded_file in files:
            OrderDocument.objects.create(
                order=order,
                document_type=document_type,
                file=uploaded_file,
                original_name=uploaded_file.name,
                file_size=uploaded_file.size,
                uploaded_by=request.user,
            )
    WorkflowEvent.objects.create(
        order=order,
        actor=request.user,
        event_type=WorkflowEvent.EventType.STATUS,
        to_status=str(order.order_status),
        note='销售下单，进入商务技术评审。',
    )
    return JsonResponse({'ok': True, 'order': _order_payload(order)}, json_dumps_params={'ensure_ascii': False})


def download_order_document(request, document_id):
    if not request.user.is_authenticated:
        return JsonResponse({'ok': False, 'error': '请先登录'}, status=401, json_dumps_params={'ensure_ascii': False})

    document = get_object_or_404(OrderDocument.objects.select_related('order'), pk=document_id)
    if not _orders_for_user(request.user).filter(pk=document.order_id).exists():
        return JsonResponse({'ok': False, 'error': '无权下载该订单文件'}, status=403, json_dumps_params={'ensure_ascii': False})
    if not document.file or not document.file.storage.exists(document.file.name):
        return JsonResponse({'ok': False, 'error': '文件不存在或已被移除'}, status=404, json_dumps_params={'ensure_ascii': False})

    return FileResponse(
        document.file.open('rb'),
        as_attachment=True,
        filename=document.original_name,
    )


def view_sample_photo(request, photo_id):
    if not request.user.is_authenticated:
        return JsonResponse({'ok': False, 'error': '请先登录'}, status=401, json_dumps_params={'ensure_ascii': False})
    photo = get_object_or_404(SamplePhoto.objects.select_related('order'), pk=photo_id)
    can_view = _orders_for_user(request.user).filter(pk=photo.order_id).exists()
    lab_type = _user_lab_type(request.user)
    if not can_view and lab_type:
        can_view = SchedulePlan.objects.filter(pk=photo.schedule_id).filter(_lab_schedule_query(lab_type)).exists()
    if not can_view:
        return JsonResponse({'ok': False, 'error': '无权查看该样品照片'}, status=403, json_dumps_params={'ensure_ascii': False})
    if not photo.file or not photo.file.storage.exists(photo.file.name):
        return JsonResponse({'ok': False, 'error': '样品照片不存在或已被移除'}, status=404, json_dumps_params={'ensure_ascii': False})
    content_type = mimetypes.guess_type(photo.original_name)[0] or 'application/octet-stream'
    return FileResponse(photo.file.open('rb'), content_type=content_type, filename=photo.original_name)


def download_test_report(request, report_id):
    if not request.user.is_authenticated:
        return JsonResponse({'ok': False, 'error': '请先登录'}, status=401, json_dumps_params={'ensure_ascii': False})
    report = get_object_or_404(TestReport.objects.select_related('order'), pk=report_id)
    can_view = _orders_for_user(request.user).filter(pk=report.order_id).exists()
    lab_type = _user_lab_type(request.user)
    if not can_view and lab_type:
        can_view = report.order.schedules.filter(_lab_schedule_query(lab_type)).exists()
    if not can_view:
        return JsonResponse({'ok': False, 'error': '无权下载该检测报告'}, status=403, json_dumps_params={'ensure_ascii': False})
    if not report.report_file or not report.report_file.storage.exists(report.report_file.name):
        return JsonResponse({'ok': False, 'error': '报告 PDF 尚未生成或已被移除'}, status=404, json_dumps_params={'ensure_ascii': False})
    filename = f'{report.report_no}-{report.get_report_type_display()}.pdf'
    return FileResponse(report.report_file.open('rb'), content_type='application/pdf', as_attachment=True, filename=filename)


@csrf_exempt
@transaction.atomic
def lims_action(request):
    if request.method != 'POST':
        return HttpResponseNotAllowed(['POST'])
    auth_error = _require_auth(request)
    if auth_error:
        return auth_error

    payload, parse_error = _json_payload(request)
    if parse_error:
        return parse_error

    action = (payload.get('action') or '').strip()
    handlers = {
        'review_pass': _action_review_pass,
        'review_reject': _action_review_reject,
        'order_update': _action_order_update,
        'order_cancel': _action_order_cancel,
        'sales_confirm': _action_sales_confirm,
        'create_change': _action_create_change,
        'schedule_assign': _action_schedule_assign,
        'process_change': _action_process_change,
        'start_test': _action_start_test,
        'end_test': _action_end_test,
        'outsource_result': _action_outsource_result,
        'submit_test': _action_submit_test,
        'sample_outbound': _action_sample_outbound,
        'issue_report': _action_issue_report,
        'report_sales_pass': _action_report_sales_pass,
        'report_sales_reject': _action_report_sales_reject,
        'report_gm_pass': _action_report_gm_pass,
        'report_gm_reject': _action_report_gm_reject,
        'preinvoice_create': _action_preinvoice_create,
        'invoice_create': _action_invoice_create,
        'invoice_pay': _action_invoice_pay,
        'standard_create': _action_standard_create,
    }
    handler = handlers.get(action)
    if not handler:
        return JsonResponse({'ok': False, 'error': '未知流程动作'}, status=400, json_dumps_params={'ensure_ascii': False})
    return handler(request, payload)


def _action_review_pass(request, payload):
    role_error = _require_role(request.user, ROLE_BUSINESS, ROLE_TECH)
    if role_error:
        return role_error
    order, error = _get_order(payload)
    if error:
        return error
    if order.order_status != LabOrder.Status.PENDING_REVIEW:
        return JsonResponse({'ok': False, 'error': '只有待评审订单可以评审通过'}, status=400, json_dumps_params={'ensure_ascii': False})
    roles = set(_roles(request.user))
    is_business = ROLE_BUSINESS in roles
    is_tech = ROLE_TECH in roles
    if is_tech and order.workflow_version == LabOrder.WorkflowVersion.LAB_DIRECT:
        routing_error = _configure_v2_routes(order, request.user, payload)
        if routing_error:
            return JsonResponse({'ok': False, 'error': routing_error}, status=400, json_dumps_params={'ensure_ascii': False})
    BusinessReview.objects.create(
        order=order,
        biz_review_user=request.user if is_business else None,
        tech_review_user=request.user if is_tech else None,
        biz_quote_detail=payload.get('biz_quote_detail') or '商务技术联合评审通过。',
        tech_feasible=True,
        review_result=True,
        review_time=timezone.now(),
    )
    has_business_pass = order.reviews.filter(review_result=True, biz_review_user__isnull=False).exists()
    has_tech_pass = order.reviews.filter(review_result=True, tech_review_user__isnull=False).exists()
    if has_business_pass and has_tech_pass:
        if order.workflow_version == LabOrder.WorkflowVersion.LAB_DIRECT:
            order.mark_status(LabOrder.Status.SCHEDULING, request.user, '商务与技术双评审通过，技术已分配至实验室负责人')
            return _status_response('双评审已通过，订单已直接进入实验室负责人工作台', order)
        order.mark_status(LabOrder.Status.SCHEDULING, request.user, '商务与技术双评审均已通过，进入质量部排期')
        return _status_response('商务与技术均已评审通过，订单已进入质量部排期', order)
    waiting_for = '技术评审' if has_business_pass else '商务评审'
    _event(order, request.user, f'当前评审通过，等待{waiting_for}完成后进入排期', event_type=WorkflowEvent.EventType.REVIEW)
    return _status_response(f'当前评审已通过，订单仍在待评审，等待{waiting_for}', order)


def _action_review_reject(request, payload):
    role_error = _require_role(request.user, ROLE_BUSINESS, ROLE_TECH)
    if role_error:
        return role_error
    order, error = _get_order(payload)
    if error:
        return error
    if order.order_status != LabOrder.Status.PENDING_REVIEW:
        return JsonResponse({'ok': False, 'error': '只有待评审订单可以驳回'}, status=400, json_dumps_params={'ensure_ascii': False})
    reason = payload.get('reject_reason') or '评审不通过，退回销售补充信息。'
    BusinessReview.objects.create(
        order=order,
        biz_review_user=request.user if ROLE_BUSINESS in _roles(request.user) else _first_user_in_group(ROLE_BUSINESS),
        tech_review_user=request.user if ROLE_TECH in _roles(request.user) else _first_user_in_group(ROLE_TECH),
        biz_quote_detail=payload.get('biz_quote_detail') or '',
        tech_feasible=bool(payload.get('tech_feasible', False)),
        review_result=False,
        reject_reason=reason,
        review_time=timezone.now(),
    )
    if order.workflow_version == LabOrder.WorkflowVersion.LAB_DIRECT:
        order.schedules.filter(samples__isnull=True, experiments__isnull=True).delete()
        order.lead_lab_manager = None
        order.save(update_fields=['lead_lab_manager', 'update_time'])
    order.mark_status(LabOrder.Status.REVIEW_REJECTED, request.user, f'商务技术评审驳回：{reason}')
    return _status_response('评审已驳回，订单回到销售', order)


def _action_order_update(request, payload):
    role_error = _require_role(request.user, ROLE_SALES)
    if role_error:
        return role_error
    order, error = _get_order(payload)
    if error:
        return error
    if not _is_chairman(request.user) and order.sale_user_id != request.user.id:
        return JsonResponse({'ok': False, 'error': '销售只能修改自己的订单'}, status=403, json_dumps_params={'ensure_ascii': False})
    if order.order_status not in [LabOrder.Status.PENDING_REVIEW, LabOrder.Status.REVIEW_REJECTED]:
        return JsonResponse({'ok': False, 'error': '当前订单状态不可修改'}, status=400, json_dumps_params={'ensure_ascii': False})
    order.customer_name = payload.get('customer_name') or order.customer_name
    order.customer_contact = payload.get('contact_name') or order.customer_contact
    order.customer_phone = payload.get('phone') or order.customer_phone
    order.project_name = payload.get('project_name') or order.project_name
    order.test_demand = payload.get('test_demand') or payload.get('test_requirements') or order.test_demand
    if payload.get('test_method') is not None:
        order.test_method = str(payload.get('test_method')).strip()
    if payload.get('test_standard') is not None:
        order.test_standard = str(payload.get('test_standard')).strip()
    if payload.get('quoted_amount') not in [None, '']:
        order.total_quote = Decimal(str(payload.get('quoted_amount')))
    order.expect_sample_arrive = _parse_datetime(payload.get('expected_sample_arrival')) or order.expect_sample_arrive
    order.expect_delivery_time = _parse_datetime(payload.get('expected_delivery_date')) or order.expect_delivery_time
    order.sales_confirmed_at = None
    if order.workflow_version == LabOrder.WorkflowVersion.LAB_DIRECT:
        order.lead_lab_manager = None
        order.schedules.filter(samples__isnull=True, experiments__isnull=True).delete()
    order.order_status = LabOrder.Status.PENDING_REVIEW
    order.update_by = request.user.username
    order.save()
    _event(order, request.user, '销售修改订单后重新提交商务技术评审')
    return _status_response('订单已修改并重新提交评审', order)


def _action_order_cancel(request, payload):
    role_error = _require_role(request.user, ROLE_SALES)
    if role_error:
        return role_error
    order, error = _get_order(payload)
    if error:
        return error
    if not _is_chairman(request.user) and order.sale_user_id != request.user.id:
        return JsonResponse({'ok': False, 'error': '销售只能退自己的订单'}, status=403, json_dumps_params={'ensure_ascii': False})
    if order.order_status not in [LabOrder.Status.PENDING_REVIEW, LabOrder.Status.REVIEW_REJECTED]:
        return JsonResponse({'ok': False, 'error': '当前订单状态不可退单'}, status=400, json_dumps_params={'ensure_ascii': False})
    order.mark_status(LabOrder.Status.CANCELLED, request.user, payload.get('reason') or '销售退单，流程终止')
    return _status_response('订单已退单', order)


def _action_sales_confirm(request, payload):
    role_error = _require_role(request.user, ROLE_SALES)
    if role_error:
        return role_error
    order, error = _get_order(payload)
    if error:
        return error
    if not _is_chairman(request.user) and order.sale_user_id != request.user.id:
        return JsonResponse({'ok': False, 'error': '销售只能确认自己的订单'}, status=403, json_dumps_params={'ensure_ascii': False})
    if order.order_status != LabOrder.Status.SCHEDULING:
        return JsonResponse({'ok': False, 'error': '只有排期中订单可以确认需求'}, status=400, json_dumps_params={'ensure_ascii': False})
    order.sales_confirmed_at = timezone.now()
    order.save(update_fields=['sales_confirmed_at', 'update_time'])
    target = '实验室负责人确认到样状态并执行试验' if order.workflow_version == LabOrder.WorkflowVersion.LAB_DIRECT else '历史质量流程继续处理'
    _event(order, request.user, payload.get('note') or f'销售确认样品与需求无变更，流转{target}')
    return _status_response('销售已确认无变更', order)


def _action_create_change(request, payload):
    role_error = _require_role(request.user, ROLE_SALES, ROLE_QUALITY, ROLE_SUZHOU_LAB, ROLE_JIANGYIN_LAB, ROLE_LAB_OPERATOR)
    if role_error:
        return role_error
    order, error = _get_order(payload)
    if error:
        return error
    roles = set(_roles(request.user))
    if (
        order.workflow_version == LabOrder.WorkflowVersion.LAB_DIRECT
        and ROLE_QUALITY in roles
        and not ({ROLE_SALES, ROLE_SUZHOU_LAB, ROLE_JIANGYIN_LAB, ROLE_LAB_OPERATOR} & roles)
        and not _is_chairman(request.user)
    ):
        return JsonResponse({'ok': False, 'error': 'V2 订单已取消质量部操作权限'}, status=403, json_dumps_params={'ensure_ascii': False})
    scene = int(payload.get('change_scene') or ChangeRequest.Scene.BEFORE_SAMPLE)
    if order.workflow_version == LabOrder.WorkflowVersion.LAB_DIRECT:
        if ROLE_SALES in roles or _is_chairman(request.user):
            target_schedules = list(order.schedules.all())
        else:
            assigned_schedule = _schedule_for_actor(order, payload, request.user)
            target_schedules = [assigned_schedule] if assigned_schedule else []
    else:
        target_schedules = [order.schedules.first()] if order.schedules.exists() else []
    if not target_schedules:
        return JsonResponse({'ok': False, 'error': '没有可回流的排期任务'}, status=400, json_dumps_params={'ensure_ascii': False})

    change_content = payload.get('change_content') or '订单需求发生变更，回流负责人重新调整排期。'
    for schedule in target_schedules:
        schedule.schedule_status = SchedulePlan.Status.CHANGE_PENDING
        schedule.save(update_fields=['schedule_status', 'update_time'])
        ChangeRequest.objects.create(
            order=order,
            schedule=schedule,
            change_scene=scene,
            old_test_demand=order.test_demand,
            new_test_demand=payload.get('new_test_demand') or order.test_demand,
            change_content=change_content,
            change_user=request.user,
            change_time=timezone.now(),
            change_status=ChangeRequest.Status.PENDING,
        )
        if ROLE_SUZHOU_LAB in roles or ROLE_JIANGYIN_LAB in roles or ROLE_LAB_OPERATOR in roles:
            _event(
                order,
                request.user,
                '实验室人员发起试验变更',
                event_type=WorkflowEvent.EventType.CHANGE,
                action_code='lab_change_create',
                changes={
                    'test_demand': _audit_change('试验需求', order.test_demand, payload.get('new_test_demand') or order.test_demand),
                    'change_content': _audit_change('变更说明', '', change_content),
                },
                schedule=schedule,
            )
    order.sales_confirmed_at = None
    order.save(update_fields=['sales_confirmed_at', 'update_time'])
    order.mark_status(LabOrder.Status.SCHEDULING, request.user, f'创建变更单：{change_content}')
    return _status_response('变更单已创建，回流排期负责人', order)


SAMPLE_PHOTO_EXTENSIONS = {'.jpg', '.jpeg', '.png'}
SAMPLE_PHOTO_MAX_SIZE = 10 * 1024 * 1024
SAMPLE_PHOTO_TOTAL_MAX_SIZE = 30 * 1024 * 1024


def _update_sample_arrival(request, payload, order, schedule):
    before_arrived = schedule.sample_arrived
    arrived = _truthy(payload.get('sample_arrived'))
    photos = request.FILES.getlist('sample_photos')
    existing_photo_count = schedule.sample_photos.count()

    if arrived and not photos and not existing_photo_count:
        return JsonResponse(
            {'ok': False, 'error': '选择“样品已到”时必须上传至少一张样品照片'},
            status=400,
            json_dumps_params={'ensure_ascii': False},
        ), None
    if not arrived and schedule.experiments.exists():
        return JsonResponse(
            {'ok': False, 'error': '当前任务已经产生试验记录，不能改为“样品未到”'},
            status=400,
            json_dumps_params={'ensure_ascii': False},
        ), None
    if sum(photo.size for photo in photos) > SAMPLE_PHOTO_TOTAL_MAX_SIZE:
        return JsonResponse(
            {'ok': False, 'error': '本次上传的样品照片合计不能超过 30MB'},
            status=400,
            json_dumps_params={'ensure_ascii': False},
        ), None
    for photo in photos:
        if Path(photo.name).suffix.lower() not in SAMPLE_PHOTO_EXTENSIONS:
            return JsonResponse(
                {'ok': False, 'error': f'{photo.name} 格式不支持，仅允许 JPG、PNG'},
                status=400,
                json_dumps_params={'ensure_ascii': False},
            ), None
        if photo.content_type not in {'image/jpeg', 'image/png'}:
            return JsonResponse(
                {'ok': False, 'error': f'{photo.name} 不是有效的 JPG 或 PNG 图片类型'},
                status=400,
                json_dumps_params={'ensure_ascii': False},
            ), None
        if photo.size > SAMPLE_PHOTO_MAX_SIZE:
            return JsonResponse(
                {'ok': False, 'error': f'{photo.name} 超过 10MB'},
                status=400,
                json_dumps_params={'ensure_ascii': False},
            ), None

    schedule.sample_arrived = arrived
    schedule.sample_arrived_at = schedule.sample_arrived_at or timezone.now() if arrived else None
    schedule.sample_confirmed_by = request.user
    schedule.save(update_fields=['sample_arrived', 'sample_arrived_at', 'sample_confirmed_by', 'update_time'])
    for photo in photos:
        SamplePhoto.objects.create(
            order=order,
            schedule=schedule,
            file=photo,
            original_name=photo.name,
            file_size=photo.size,
            uploaded_by=request.user,
        )
    if arrived:
        sample, _ = Sample.objects.get_or_create(
            order=order,
            schedule=schedule,
            defaults={
                'sample_no': _next_sample_no(order),
                'sample_name': f'{order.project_name} 样品',
                'sample_spec': '客户送检样品',
                'sample_count': 1,
                'storage_condition': '按试验要求存放',
                'actual_arrive_time': schedule.sample_arrived_at,
                'sample_status': Sample.Status.REGISTERED,
                'quality_user': request.user,
            },
        )
        changed_fields = []
        if not sample.actual_arrive_time:
            sample.actual_arrive_time = schedule.sample_arrived_at
            changed_fields.append('actual_arrive_time')
        if not sample.quality_user_id:
            sample.quality_user = request.user
            changed_fields.append('quality_user')
        if changed_fields:
            sample.save(update_fields=[*changed_fields, 'update_time'])
    return None, {
        'sample_arrived': _audit_change('到样状态', '样品已到' if before_arrived else '样品未到', '样品已到' if arrived else '样品未到'),
        'sample_photos': _audit_change('新增样品照片', 0, len(photos)),
    }


def _action_schedule_assign(request, payload):
    order, error = _get_order(payload)
    if error:
        return error
    if order.workflow_version == LabOrder.WorkflowVersion.LAB_DIRECT:
        role_error = _require_role(request.user, ROLE_SUZHOU_LAB, ROLE_JIANGYIN_LAB, ROLE_LAB_OPERATOR)
    else:
        role_error = _require_role(request.user, ROLE_QUALITY)
    if role_error:
        return role_error
    if order.order_status not in [LabOrder.Status.SCHEDULING, LabOrder.Status.TESTING]:
        return JsonResponse({'ok': False, 'error': '当前订单不可排期'}, status=400, json_dumps_params={'ensure_ascii': False})
    if order.workflow_version == LabOrder.WorkflowVersion.LAB_DIRECT:
        schedule = _schedule_for_actor(order, payload, request.user)
        if not schedule:
            return JsonResponse({'ok': False, 'error': '没有分配给当前负责人的任务'}, status=403, json_dumps_params={'ensure_ascii': False})
        if schedule.schedule_status in [SchedulePlan.Status.ENDED, SchedulePlan.Status.FINISHED]:
            return JsonResponse(
                {'ok': False, 'error': '实验已结束，不能重新排期或更换设备'},
                status=400,
                json_dumps_params={'ensure_ascii': False},
            )
        before_start = schedule.plan_start_time
        before_end = schedule.plan_end_time
        before_device = schedule.device.device_name if schedule.device else ''
        start_time, end_time, range_error = _parse_plan_range(
            payload.get('plan_start_time'), payload.get('plan_end_time')
        )
        if range_error:
            return JsonResponse({'ok': False, 'error': range_error}, status=400, json_dumps_params={'ensure_ascii': False})
        device_error = _assign_schedule_device(schedule, payload, start_time, end_time)
        if device_error:
            return JsonResponse({'ok': False, 'error': device_error}, status=400, json_dumps_params={'ensure_ascii': False})
        schedule.plan_start_time = start_time
        schedule.plan_end_time = end_time
        if schedule.test_type == SchedulePlan.TestType.OUTSOURCE:
            schedule.outsource_factory = payload.get('outsource_factory') or schedule.outsource_factory
            schedule.outsource_price = Decimal(str(payload.get('outsource_price') or schedule.outsource_price or '0'))
            schedule.outsource_cycle = int(payload.get('outsource_cycle') or schedule.outsource_cycle or 0) or None
            if not schedule.outsource_factory:
                return JsonResponse({'ok': False, 'error': '委外任务必须填写委外厂家'}, status=400, json_dumps_params={'ensure_ascii': False})
        schedule.quality_user = request.user
        schedule.remark = payload.get('remark') or schedule.remark or order.test_demand
        schedule.save()
        sample_error, sample_changes = _update_sample_arrival(request, payload, order, schedule)
        if sample_error:
            transaction.set_rollback(True)
            return sample_error
        _event(
            order,
            request.user,
            f'{schedule.get_test_type_display()}完成任务排期与排台',
            action_code='lab_schedule_assign',
            changes={
                'plan_start_time': _audit_change('计划开始', before_start, schedule.plan_start_time),
                'plan_end_time': _audit_change('计划结束', before_end, schedule.plan_end_time),
                'device': _audit_change('试验设备', before_device, schedule.device.device_name if schedule.device else ''),
                'outsource_factory': _audit_change('委外厂家', '', schedule.outsource_factory),
                **sample_changes,
            },
            schedule=schedule,
        )
        return _status_response('当前实验室任务排期已更新', order)

    test_type = int(payload.get('test_type') or SchedulePlan.TestType.SUZHOU)
    manager = None
    if test_type == SchedulePlan.TestType.SUZHOU:
        manager = _first_user_in_group(ROLE_SUZHOU_LAB)
    elif test_type == SchedulePlan.TestType.JIANGYIN:
        manager = _first_user_in_group(ROLE_JIANGYIN_LAB)
    SchedulePlan.objects.create(
        order=order,
        test_type=test_type,
        lab_manager=manager,
        outsource_factory=payload.get('outsource_factory') or '',
        outsource_price=Decimal(str(payload.get('outsource_price') or '0')),
        outsource_cycle=int(payload.get('outsource_cycle') or 0) or None,
        plan_start_time=_parse_datetime(payload.get('plan_start_time')) or timezone.now(),
        plan_end_time=_parse_datetime(payload.get('plan_end_time')) or (timezone.now() + timezone.timedelta(days=7)),
        schedule_status=SchedulePlan.Status.NEW,
        quality_user=request.user,
        remark=payload.get('remark') or '质量部排期分配',
    )
    order.order_status = LabOrder.Status.SCHEDULING
    order.execution_mode = test_type if test_type in [1, 2, 3] else order.execution_mode
    order.update_by = request.user.username
    order.save()
    _event(order, request.user, '质量部完成试验排期分配，生成项目周期表')
    return _status_response('排期已创建', order)


def _action_process_change(request, payload):
    order, error = _get_order(payload)
    if error:
        return error
    if order.workflow_version == LabOrder.WorkflowVersion.LAB_DIRECT:
        role_error = _require_role(request.user, ROLE_SUZHOU_LAB, ROLE_JIANGYIN_LAB, ROLE_LAB_OPERATOR)
    else:
        role_error = _require_role(request.user, ROLE_QUALITY)
    if role_error:
        return role_error
    pending_changes = order.change_requests.exclude(change_status=ChangeRequest.Status.APPLIED)
    if payload.get('schedule_id'):
        pending_changes = pending_changes.filter(schedule_id=payload.get('schedule_id'))
    if order.workflow_version == LabOrder.WorkflowVersion.LAB_DIRECT and not _is_chairman(request.user):
        if _is_lab_operator(request.user):
            lab_type = _user_lab_type(request.user)
            pending_changes = pending_changes.filter(schedule__in=order.schedules.filter(_lab_schedule_query(lab_type)))
        else:
            pending_changes = pending_changes.filter(schedule__lab_manager=request.user)
    change = pending_changes.first()
    if not change:
        return JsonResponse({'ok': False, 'error': '没有待处理变更单'}, status=400, json_dumps_params={'ensure_ascii': False})
    before_start = change.schedule.plan_start_time if change.schedule else None
    before_end = change.schedule.plan_end_time if change.schedule else None
    before_device = change.schedule.device.device_name if change.schedule and change.schedule.device else ''
    if change.schedule:
        start_time, end_time, range_error = _parse_plan_range(
            payload.get('plan_start_time'), payload.get('plan_end_time')
        )
        if range_error:
            return JsonResponse({'ok': False, 'error': range_error}, status=400, json_dumps_params={'ensure_ascii': False})
        device_error = _assign_schedule_device(change.schedule, payload, start_time, end_time)
        if device_error:
            return JsonResponse({'ok': False, 'error': device_error}, status=400, json_dumps_params={'ensure_ascii': False})
        change.schedule.plan_start_time = start_time
        change.schedule.plan_end_time = end_time
        change.schedule.schedule_status = SchedulePlan.Status.NEW
        change.schedule.save()
        sample_error, sample_changes = _update_sample_arrival(request, payload, order, change.schedule)
        if sample_error:
            transaction.set_rollback(True)
            return sample_error
    change.change_status = ChangeRequest.Status.APPLIED
    change.save(update_fields=['change_status', 'update_time'])
    _event(
        order,
        request.user,
        '实验室人员已处理变更单并更新项目周期表',
        event_type=WorkflowEvent.EventType.CHANGE,
        action_code='lab_change_process',
        changes={
            'plan_start_time': _audit_change('计划开始', before_start, change.schedule.plan_start_time if change.schedule else None),
            'plan_end_time': _audit_change('计划结束', before_end, change.schedule.plan_end_time if change.schedule else None),
            'device': _audit_change('试验设备', before_device, change.schedule.device.device_name if change.schedule and change.schedule.device else ''),
            'change_status': _audit_change('变更状态', '待调整', '已闭环'),
            **(sample_changes or {}),
        },
        schedule=change.schedule,
    )
    return _status_response('变更已闭环', order)


def _action_start_test(request, payload):
    role_error = _require_role(request.user, ROLE_SUZHOU_LAB, ROLE_JIANGYIN_LAB, ROLE_LAB_OPERATOR)
    if role_error:
        return role_error
    order, error = _get_order(payload)
    if error:
        return error
    schedule = _schedule_for_actor(order, payload, request.user)
    if not schedule:
        return JsonResponse({'ok': False, 'error': '没有分配给当前实验室负责人的排期'}, status=403, json_dumps_params={'ensure_ascii': False})
    if schedule.schedule_status in [SchedulePlan.Status.ENDED, SchedulePlan.Status.FINISHED]:
        return JsonResponse({'ok': False, 'error': '该试验任务已经结束，不能重复开始'}, status=400, json_dumps_params={'ensure_ascii': False})
    if order.workflow_version == LabOrder.WorkflowVersion.LAB_DIRECT and not order.sales_confirmed_at:
        return JsonResponse({'ok': False, 'error': '销售尚未确认需求，不能开始试验'}, status=400, json_dumps_params={'ensure_ascii': False})
    if order.workflow_version == LabOrder.WorkflowVersion.LAB_DIRECT:
        if not schedule.device:
            return JsonResponse({'ok': False, 'error': '请先完成设备排台'}, status=400, json_dumps_params={'ensure_ascii': False})
        if schedule.device.device_status != LabDevice.Status.NORMAL:
            return JsonResponse(
                {'ok': False, 'error': f'设备当前为“{schedule.device.get_device_status_display()}”，不能开始试验'},
                status=400,
                json_dumps_params={'ensure_ascii': False},
            )
    if not schedule.sample_arrived:
        return JsonResponse({'ok': False, 'error': '样品尚未到达，不能开始试验'}, status=400, json_dumps_params={'ensure_ascii': False})
    existing_experiment = order.experiments.filter(schedule=schedule).order_by('-create_time').first()
    if existing_experiment and existing_experiment.test_status in [Experiment.Status.ENDED, Experiment.Status.FINISHED]:
        return JsonResponse(
            {'ok': False, 'error': '该实验已结束，不能重复开始'},
            status=400,
            json_dumps_params={'ensure_ascii': False},
        )
    sample = order.samples.filter(schedule=schedule).first()
    if sample:
        sample.sample_status = Sample.Status.TESTING
        sample.save(update_fields=['sample_status', 'update_time'])
    schedule.schedule_status = SchedulePlan.Status.RUNNING
    schedule.save(update_fields=['schedule_status', 'update_time'])
    experiment, _ = Experiment.objects.get_or_create(
        order=order,
        schedule=schedule,
        sample=sample,
        defaults={
            'test_item_list': schedule.remark or payload.get('test_item_list') or order.test_demand,
            'test_standard': order.test_standard or '订单未指定',
            'test_start_time': timezone.now(),
            'test_operator': request.user,
            'test_status': Experiment.Status.RUNNING,
            'test_type': schedule.test_type,
        },
    )
    experiment.test_item_list = schedule.remark or payload.get('test_item_list') or order.test_demand
    experiment.test_standard = order.test_standard or '订单未指定'
    experiment.test_start_time = experiment.test_start_time or timezone.now()
    experiment.test_operator = request.user
    experiment.test_status = Experiment.Status.RUNNING
    experiment.test_type = schedule.test_type
    experiment.save()
    order.mark_status(LabOrder.Status.TESTING, request.user, '实验室开始试验')
    _event(
        order,
        request.user,
        '实验室人员开始试验',
        action_code='lab_test_start',
        changes={
            'test_item_list': _audit_change('试验项目', '', experiment.test_item_list),
            'test_standard': _audit_change('执行标准', '', experiment.test_standard),
            'test_status': _audit_change('试验状态', '待开展', experiment.get_test_status_display()),
            'test_operator': _audit_change('操作人员', '', _display_user(request.user)),
        },
        schedule=schedule,
    )
    return _status_response('试验已开始', order)


def _action_standard_create(request, payload):
    role_error = _require_role(request.user, ROLE_SUZHOU_LAB, ROLE_JIANGYIN_LAB, ROLE_QUALITY)
    if role_error:
        return role_error
    industry = (payload.get('industry') or '').strip()
    standard_code = (payload.get('standard_code') or '').strip()
    standard_name = (payload.get('standard_name') or '').strip()
    description = (payload.get('description') or '').strip()
    if not industry or not standard_code or not standard_name:
        return JsonResponse({'ok': False, 'error': '行业、标准编号、标准名称必填'}, status=400, json_dumps_params={'ensure_ascii': False})
    standard, created = TestStandard.objects.update_or_create(
        industry=industry,
        standard_code=standard_code,
        defaults={
            'standard_name': standard_name,
            'description': description,
            'is_active': True,
        },
    )
    return JsonResponse(
        {
            'ok': True,
            'message': '试验标准已添加' if created else '试验标准已更新',
            'standard': _standard_payload(standard),
        },
        json_dumps_params={'ensure_ascii': False},
    )


def _experiment_result_from_payload(payload):
    result_status = str(payload.get('result_status') or '').strip()
    if result_status not in Experiment.Result.values:
        return None
    return result_status


def _all_test_paths_finished(order):
    schedules = order.schedules.all()
    if not schedules.exists() or schedules.exclude(schedule_status=SchedulePlan.Status.FINISHED).exists():
        return False
    if order.experiments.exclude(test_status=Experiment.Status.FINISHED).exists():
        return False
    return not schedules.exclude(experiments__test_status=Experiment.Status.FINISHED).exists()


def _sync_order_test_completion(order, actor):
    if _all_test_paths_finished(order):
        if order.order_status != LabOrder.Status.TEST_FINISHED:
            order.mark_status(
                LabOrder.Status.TEST_FINISHED,
                actor,
                '全部执行路径的实验结果均已提交，等待主责实验室负责人出具报告',
            )
        return True

    schedules = order.schedules.all()
    all_paths_ended = (
        schedules.exists()
        and not schedules.exclude(
            schedule_status__in=[SchedulePlan.Status.ENDED, SchedulePlan.Status.FINISHED]
        ).exists()
    )
    has_unsubmitted_result = schedules.filter(schedule_status=SchedulePlan.Status.ENDED).exists()
    if all_paths_ended and has_unsubmitted_result:
        if order.order_status != LabOrder.Status.RESULT_PENDING:
            order.mark_status(
                LabOrder.Status.RESULT_PENDING,
                actor,
                '全部实验已结束，尚有实验结果待提交',
            )
        return False

    if order.order_status != LabOrder.Status.TESTING:
        order.mark_status(LabOrder.Status.TESTING, actor, '部分执行路径已结束，其他试验任务继续执行')
    return False


def _action_outsource_result(request, payload):
    order, error = _get_order(payload)
    if error:
        return error
    if order.workflow_version == LabOrder.WorkflowVersion.LAB_DIRECT:
        role_error = _require_role(request.user, ROLE_SUZHOU_LAB, ROLE_JIANGYIN_LAB, ROLE_LAB_OPERATOR)
    else:
        role_error = _require_role(request.user, ROLE_QUALITY)
    if role_error:
        return role_error
    schedules = order.schedules.filter(test_type=SchedulePlan.TestType.OUTSOURCE)
    if payload.get('schedule_id'):
        schedules = schedules.filter(id=payload.get('schedule_id'))
    if order.workflow_version == LabOrder.WorkflowVersion.LAB_DIRECT and not _is_chairman(request.user):
        schedule = next((item for item in schedules if _can_operate_schedule(request.user, item)), None)
    else:
        schedule = schedules.order_by('-create_time').first()
    if not schedule:
        return JsonResponse({'ok': False, 'error': '该订单没有委外排期'}, status=400, json_dumps_params={'ensure_ascii': False})
    if order.workflow_version == LabOrder.WorkflowVersion.LAB_DIRECT and not order.sales_confirmed_at:
        return JsonResponse({'ok': False, 'error': '销售尚未确认需求，不能回传委外试验结果'}, status=400, json_dumps_params={'ensure_ascii': False})
    if not schedule.sample_arrived:
        return JsonResponse({'ok': False, 'error': '委外样品尚未到达，不能回传试验结果'}, status=400, json_dumps_params={'ensure_ascii': False})
    existing_experiment = order.experiments.filter(schedule=schedule).order_by('-create_time').first()
    if existing_experiment and existing_experiment.test_status == Experiment.Status.ENDED:
        return JsonResponse(
            {'ok': False, 'error': '委外实验已结束并保存结果，请直接点击“提交结果”'},
            status=400,
            json_dumps_params={'ensure_ascii': False},
        )
    if existing_experiment and existing_experiment.test_status == Experiment.Status.FINISHED:
        return JsonResponse(
            {'ok': False, 'error': '委外实验结果已经提交，不能重复回传'},
            status=400,
            json_dumps_params={'ensure_ascii': False},
        )
    result_status = _experiment_result_from_payload(payload)
    if not result_status:
        return JsonResponse({'ok': False, 'error': '请选择实验结果'}, status=400, json_dumps_params={'ensure_ascii': False})
    sample = order.samples.filter(schedule=schedule).first() or order.samples.first()
    experiment, _ = Experiment.objects.get_or_create(
        order=order,
        schedule=schedule,
        sample=sample,
        defaults={
            'test_item_list': payload.get('test_item_list') or schedule.remark or order.test_demand,
            'test_standard': payload.get('test_standard') or '委外厂家回传标准',
            'test_start_time': _parse_datetime(payload.get('test_start_time')) or schedule.plan_start_time or timezone.now(),
            'test_type': SchedulePlan.TestType.OUTSOURCE,
        },
    )
    experiment.test_item_list = payload.get('test_item_list') or experiment.test_item_list or schedule.remark or order.test_demand
    experiment.test_standard = payload.get('test_standard') or experiment.test_standard or '委外厂家回传标准'
    experiment.test_raw_data = payload.get('test_raw_data') or experiment.test_raw_data or '委外厂家已回传原始试验数据'
    experiment.test_conclusion_temp = payload.get('test_conclusion_temp') or experiment.test_conclusion_temp or '委外试验完成，等待主责实验室出具报告'
    experiment.result_status = result_status
    experiment.test_start_time = experiment.test_start_time or schedule.plan_start_time or timezone.now()
    experiment.test_end_time = _parse_datetime(payload.get('test_end_time')) or timezone.now()
    experiment.test_operator = request.user
    experiment.test_status = Experiment.Status.ENDED
    experiment.test_type = SchedulePlan.TestType.OUTSOURCE
    experiment.save()
    if sample:
        sample.sample_status = Sample.Status.FINISHED
        sample.save(update_fields=['sample_status', 'update_time'])
    schedule.schedule_status = SchedulePlan.Status.ENDED
    schedule.save(update_fields=['schedule_status', 'update_time'])
    _event(
        order,
        request.user,
        '实验室人员录入委外试验结果，等待正式提交',
        action_code='lab_outsource_result_recorded',
        changes={
            'test_status': _audit_change('委外试验状态', '待回传', experiment.get_test_status_display()),
            'test_end_time': _audit_change('委外完成时间', '', experiment.test_end_time),
            'test_raw_data': _audit_change('回传数据摘要', '', experiment.test_raw_data[:500]),
            'test_conclusion': _audit_change('委外结论', '', experiment.test_conclusion_temp),
            'result_status': _audit_change('实验结果', '', experiment.get_result_status_display()),
        },
        schedule=schedule,
    )
    _sync_order_test_completion(order, request.user)
    return _status_response('委外实验已结束并记录结果，请确认后点击“提交结果”', order)


def _action_end_test(request, payload):
    role_error = _require_role(request.user, ROLE_SUZHOU_LAB, ROLE_JIANGYIN_LAB, ROLE_LAB_OPERATOR)
    if role_error:
        return role_error
    order, error = _get_order(payload)
    if error:
        return error
    schedule = _schedule_for_actor(order, payload, request.user)
    experiment = order.experiments.filter(schedule=schedule, test_status=Experiment.Status.RUNNING).first() if schedule else None
    if not experiment:
        return JsonResponse({'ok': False, 'error': '没有正在进行的实验记录'}, status=400, json_dumps_params={'ensure_ascii': False})
    result_status = _experiment_result_from_payload(payload)
    if not result_status:
        return JsonResponse({'ok': False, 'error': '请选择实验结果'}, status=400, json_dumps_params={'ensure_ascii': False})
    experiment.test_raw_data = payload.get('test_raw_data') or experiment.test_raw_data or '试验数据已录入。'
    experiment.test_conclusion_temp = payload.get('test_conclusion_temp') or '试验完成，等待主责实验室出报告。'
    experiment.result_status = result_status
    experiment.test_end_time = timezone.now()
    experiment.test_status = Experiment.Status.ENDED
    experiment.save()
    if experiment.sample:
        experiment.sample.sample_status = Sample.Status.FINISHED
        experiment.sample.save(update_fields=['sample_status', 'update_time'])
    if experiment.schedule:
        experiment.schedule.schedule_status = SchedulePlan.Status.ENDED
        experiment.schedule.save(update_fields=['schedule_status', 'update_time'])
    _event(
        order,
        request.user,
        '实验室人员结束实验并记录结果',
        action_code='lab_test_end',
        changes={
            'test_raw_data': _audit_change('原始数据摘要', '', experiment.test_raw_data[:500]),
            'test_conclusion': _audit_change('试验结论', '', experiment.test_conclusion_temp),
            'result_status': _audit_change('实验结果', '', experiment.get_result_status_display()),
            'test_status': _audit_change('试验状态', '试验中', experiment.get_test_status_display()),
            'test_end_time': _audit_change('实际结束时间', '', experiment.test_end_time),
        },
        schedule=experiment.schedule,
    )
    _event(order, request.user, '实验已结束并保存结果，尚未触发后续报告流程')
    _sync_order_test_completion(order, request.user)
    return _status_response('实验已结束，结果已保存；请确认后点击“提交结果”', order)


def _action_submit_test(request, payload):
    role_error = _require_role(request.user, ROLE_SUZHOU_LAB, ROLE_JIANGYIN_LAB, ROLE_LAB_OPERATOR, ROLE_QUALITY)
    if role_error:
        return role_error
    order, error = _get_order(payload)
    if error:
        return error
    schedule = _schedule_for_actor(order, payload, request.user)
    if not schedule:
        return JsonResponse({'ok': False, 'error': '没有分配给当前人员的实验任务'}, status=403, json_dumps_params={'ensure_ascii': False})
    experiment = order.experiments.filter(schedule=schedule).order_by('-create_time').first()
    if not experiment:
        return JsonResponse({'ok': False, 'error': '请先开始并结束实验'}, status=400, json_dumps_params={'ensure_ascii': False})
    if experiment.test_status == Experiment.Status.FINISHED:
        return JsonResponse({'ok': False, 'error': '该实验结果已经提交'}, status=400, json_dumps_params={'ensure_ascii': False})
    if experiment.test_status != Experiment.Status.ENDED:
        return JsonResponse({'ok': False, 'error': '实验尚未结束，不能提交结果'}, status=400, json_dumps_params={'ensure_ascii': False})
    if not experiment.result_status:
        return JsonResponse({'ok': False, 'error': '实验结果不完整，请先执行“实验结束”并填写结果'}, status=400, json_dumps_params={'ensure_ascii': False})

    experiment.test_status = Experiment.Status.FINISHED
    experiment.save(update_fields=['test_status', 'update_time'])
    schedule.schedule_status = SchedulePlan.Status.FINISHED
    schedule.save(update_fields=['schedule_status', 'update_time'])
    _event(
        order,
        request.user,
        '实验室人员正式提交实验结果',
        action_code='lab_test_result_submit',
        changes={
            'test_status': _audit_change('实验状态', '实验已结束待提交结果', experiment.get_test_status_display()),
            'result_status': _audit_change('实验结果', '', experiment.get_result_status_display()),
            'test_conclusion': _audit_change('实验结论', '', experiment.test_conclusion_temp),
        },
        schedule=schedule,
    )
    all_finished = _sync_order_test_completion(order, request.user)
    message = '结果已提交；全部执行路径均已提交，可出具报告' if all_finished else '本任务结果已提交，等待其他执行路径提交结果'
    return _status_response(message, order)


def _action_sample_outbound(request, payload):
    role_error = _require_role(request.user, ROLE_SUZHOU_LAB, ROLE_JIANGYIN_LAB, ROLE_LAB_OPERATOR)
    if role_error:
        return role_error
    order, error = _get_order(payload)
    if error:
        return error
    schedule = _schedule_for_actor(order, payload, request.user)
    if not schedule or not _can_operate_schedule(request.user, schedule):
        return JsonResponse(
            {'ok': False, 'error': '没有分配给当前实验室人员的样品任务'},
            status=403,
            json_dumps_params={'ensure_ascii': False},
        )
    if not schedule.sample_arrived:
        return JsonResponse({'ok': False, 'error': '样品尚未入库，不能办理出库'}, status=400, json_dumps_params={'ensure_ascii': False})
    if schedule.schedule_status not in [SchedulePlan.Status.ENDED, SchedulePlan.Status.FINISHED]:
        return JsonResponse({'ok': False, 'error': '试验尚未完成，不能办理样品出库'}, status=400, json_dumps_params={'ensure_ascii': False})
    samples = list(schedule.samples.select_for_update().order_by('id'))
    if not samples:
        return JsonResponse({'ok': False, 'error': '当前任务没有可出库的样品记录'}, status=400, json_dumps_params={'ensure_ascii': False})
    if all(sample.outbound_time for sample in samples):
        return JsonResponse({'ok': False, 'error': '当前任务样品已经出库'}, status=400, json_dumps_params={'ensure_ascii': False})

    outbound_time = timezone.now()
    for sample in samples:
        sample.outbound_time = outbound_time
        sample.outbound_by = request.user
        sample.sample_status = Sample.Status.RETURNED
        sample.save(update_fields=['outbound_time', 'outbound_by', 'sample_status', 'update_time'])
    _event(
        order,
        request.user,
        '实验室人员办理样品出库',
        action_code='lab_sample_outbound',
        changes={
            'sample_status': _audit_change('样品状态', '试验完成', '已出库'),
            'outbound_time': _audit_change('样品出库时间', '', outbound_time),
            'outbound_by': _audit_change('出库操作人', '', _display_user(request.user)),
        },
        schedule=schedule,
    )
    return _status_response('样品出库已登记', order)


def _action_issue_report(request, payload):
    order, error = _get_order(payload)
    if error:
        return error
    if order.workflow_version == LabOrder.WorkflowVersion.LAB_DIRECT:
        role_error = _require_role(request.user, ROLE_SUZHOU_LAB, ROLE_JIANGYIN_LAB)
        if role_error:
            return role_error
        if not _is_chairman(request.user) and order.lead_lab_manager_id != request.user.id:
            return JsonResponse({'ok': False, 'error': '仅本订单主责实验室负责人可以汇总出具报告'}, status=403, json_dumps_params={'ensure_ascii': False})
        if order.schedules.exclude(schedule_status=SchedulePlan.Status.FINISHED).exists():
            return JsonResponse({'ok': False, 'error': '仍有执行路径未提交实验结果，暂不能出具总报告'}, status=400, json_dumps_params={'ensure_ascii': False})
        if order.experiments.exclude(test_status=Experiment.Status.FINISHED).exists() or not order.experiments.exists():
            return JsonResponse({'ok': False, 'error': '请先提交全部实验结果'}, status=400, json_dumps_params={'ensure_ascii': False})
        _sync_order_test_completion(order, request.user)
    else:
        role_error = _require_role(request.user, ROLE_QUALITY)
        if role_error:
            return role_error
    experiment = order.experiments.order_by('-create_time').first()
    if not experiment:
        return JsonResponse({'ok': False, 'error': '请先完成试验记录'}, status=400, json_dumps_params={'ensure_ascii': False})
    report_type = payload.get('report_type') or TestReport.ReportType.FORMAL
    if report_type not in TestReport.ReportType.values:
        return JsonResponse({'ok': False, 'error': '请选择有效的报告版本'}, status=400, json_dumps_params={'ensure_ascii': False})
    report = order.reports.filter(report_status=TestReport.Status.REJECTED).order_by('-create_time').first()
    if report:
        report.test_record = experiment
        report.report_type = report_type
        report.final_conclusion = payload.get('final_conclusion') or report.final_conclusion or '检测完成，形成正式检测报告。'
        report.report_status = TestReport.Status.SALES_REVIEW
        report.create_quality_user = request.user
    else:
        report = TestReport(
            order=order,
            test_record=experiment,
            report_no=payload.get('report_no') or _next_report_no(order),
            report_type=report_type,
            final_conclusion=payload.get('final_conclusion') or '检测完成，形成正式检测报告。',
            report_status=TestReport.Status.SALES_REVIEW,
            create_quality_user=request.user,
        )
    report.generated_at = timezone.now()
    report.save()
    experiments = order.experiments.select_related('test_operator', 'schedule').order_by('create_time')
    pdf_content = build_test_report_pdf(report, experiments)
    old_file_name = report.report_file.name if report.report_file else ''
    report_filename = f'{report.report_no}-{report.report_type}.pdf'
    report.report_file.save(report_filename, ContentFile(pdf_content), save=False)
    report.report_file_url = f'/api/reports/{report.id}/download/'
    report.save(update_fields=[
        'test_record', 'report_type', 'final_conclusion', 'report_status', 'create_quality_user',
        'generated_at', 'report_file', 'report_file_url', 'update_time',
    ])
    if old_file_name and old_file_name != report.report_file.name:
        report.report_file.storage.delete(old_file_name)
    _event(
        order,
        request.user,
        '实验室人员汇总并提交检测报告',
        action_code='lab_report_issue',
        changes={
            'report_no': _audit_change('报告编号', '', report.report_no),
            'report_type': _audit_change('报告版本', '', report.get_report_type_display()),
            'report_file': _audit_change('报告文件', '', report.report_file.name),
            'report_status': _audit_change('报告状态', '草稿/驳回', report.get_report_status_display()),
            'final_conclusion': _audit_change('最终结论', '', report.final_conclusion[:500]),
            'report_creator': _audit_change('报告提交人', '', _display_user(request.user)),
        },
        schedule=experiment.schedule,
    )
    order.mark_status(LabOrder.Status.REPORT_REVIEW, request.user, f'主责实验室负责人出具报告 {report.report_no}，提交销售初审')
    return _status_response('报告已提交销售初审', order)


def _audit_report(request, payload, expected_status, level, result, next_status, note):
    report, error = _get_report(payload)
    if error:
        return error
    if report.report_status != expected_status:
        return JsonResponse({'ok': False, 'error': '报告当前状态不可执行此审核'}, status=400, json_dumps_params={'ensure_ascii': False})
    ReportAudit.objects.create(
        report=report,
        audit_level=level,
        audit_user=request.user,
        audit_result=result,
        audit_opinion=payload.get('audit_opinion') or note,
        audit_time=timezone.now(),
    )
    report.report_status = next_status
    if result == ReportAudit.Result.REJECTED:
        report.remake_count += 1
    report.save()
    _event(report.order, request.user, note, event_type=WorkflowEvent.EventType.REVIEW)
    return JsonResponse({'ok': True, 'message': note, 'report': _report_payload(report)}, json_dumps_params={'ensure_ascii': False})


def _action_report_sales_pass(request, payload):
    role_error = _require_role(request.user, ROLE_SALES)
    if role_error:
        return role_error
    return _audit_report(request, payload, TestReport.Status.SALES_REVIEW, ReportAudit.Level.SALES, ReportAudit.Result.APPROVED, TestReport.Status.GM_REVIEW, '销售初审通过，提交总经理终审')


def _action_report_sales_reject(request, payload):
    role_error = _require_role(request.user, ROLE_SALES)
    if role_error:
        return role_error
    return _audit_report(request, payload, TestReport.Status.SALES_REVIEW, ReportAudit.Level.SALES, ReportAudit.Result.REJECTED, TestReport.Status.REJECTED, '销售初审驳回，退回质量部重制')


def _action_report_gm_pass(request, payload):
    role_error = _require_role(request.user, ROLE_GENERAL_MANAGER)
    if role_error:
        return role_error
    return _audit_report(request, payload, TestReport.Status.GM_REVIEW, ReportAudit.Level.GENERAL_MANAGER, ReportAudit.Result.APPROVED, TestReport.Status.APPROVED, '总经理终审通过，推送会计开票')


def _action_report_gm_reject(request, payload):
    role_error = _require_role(request.user, ROLE_GENERAL_MANAGER)
    if role_error:
        return role_error
    return _audit_report(request, payload, TestReport.Status.GM_REVIEW, ReportAudit.Level.GENERAL_MANAGER, ReportAudit.Result.REJECTED, TestReport.Status.REJECTED, '总经理终审驳回，退回质量部重制')


def _invoice_amount_from_payload(payload, default_amount=None):
    raw_amount = payload.get('invoice_amount')
    if raw_amount in [None, '']:
        if default_amount is None:
            return None, '请填写开票金额'
        amount = Decimal(str(default_amount)).quantize(Decimal('0.01'))
        if amount <= 0:
            return None, '订单已无剩余可开票金额'
        return amount, None
    try:
        amount = Decimal(str(raw_amount)).quantize(Decimal('0.01'))
    except (InvalidOperation, TypeError, ValueError):
        return None, '开票金额格式不正确'
    if amount <= 0:
        return None, '开票金额必须大于 0'
    return amount, None


def _invoice_payment_status(payload):
    raw_status = payload.get('pay_status')
    if raw_status in [None, '']:
        return Invoice.PayStatus.UNPAID, None
    try:
        status = int(raw_status)
    except (TypeError, ValueError):
        return None, '回款状态不正确'
    if status not in Invoice.PayStatus.values:
        return None, '回款状态不正确'
    return status, None


def _action_preinvoice_create(request, payload):
    role_error = _require_role(request.user, ROLE_ACCOUNTING)
    if role_error:
        return role_error
    order, error = _get_order(payload)
    if error:
        return error

    with transaction.atomic():
        order = LabOrder.objects.select_for_update().get(pk=order.pk)
        stage = _preinvoice_stage(order)
        if not stage:
            return JsonResponse(
                {'ok': False, 'error': '该订单尚未满足双评审通过或当前不允许预开票'},
                status=400,
                json_dumps_params={'ensure_ascii': False},
            )
        invoiced_total, remaining_amount = _invoice_amounts(order)
        amount, amount_error = _invoice_amount_from_payload(payload)
        if amount_error:
            return JsonResponse({'ok': False, 'error': amount_error}, status=400, json_dumps_params={'ensure_ascii': False})
        if amount >= remaining_amount:
            return JsonResponse(
                {
                    'ok': False,
                    'error': f'预开票后必须为最终总开票保留余额；当前可用余额为 {remaining_amount:.2f} 元',
                },
                status=400,
                json_dumps_params={'ensure_ascii': False},
            )
        pay_status, pay_error = _invoice_payment_status(payload)
        if pay_error:
            return JsonResponse({'ok': False, 'error': pay_error}, status=400, json_dumps_params={'ensure_ascii': False})
        invoice_no = (payload.get('invoice_no') or _next_invoice_no(order)).strip()
        if Invoice.objects.filter(invoice_no=invoice_no).exists():
            return JsonResponse({'ok': False, 'error': '发票号码已存在'}, status=400, json_dumps_params={'ensure_ascii': False})
        invoice = Invoice.objects.create(
            order=order,
            report=None,
            invoice_stage=stage,
            invoice_no=invoice_no,
            invoice_amount=amount,
            invoice_type=payload.get('invoice_type') or '增值税专票',
            invoice_date=_parse_datetime(payload.get('invoice_date')) or timezone.now(),
            pay_status=pay_status,
            finance_user=request.user,
            order_finish_flag=Invoice.FinishFlag.UNFINISHED,
        )
        _event(
            order,
            request.user,
            f'{invoice.get_invoice_stage_display()}：{invoice.invoice_no}，金额 {invoice.invoice_amount:.2f} 元；订单继续流转',
            action_code='finance_preinvoice_create',
            changes={
                'invoice_stage': _audit_change('开票阶段', '', invoice.get_invoice_stage_display()),
                'invoice_no': _audit_change('发票号', '', invoice.invoice_no),
                'invoice_amount': _audit_change('开票金额', invoiced_total, invoiced_total + invoice.invoice_amount),
                'remaining_amount': _audit_change('剩余可开', remaining_amount, remaining_amount - invoice.invoice_amount),
            },
        )
    return JsonResponse({'ok': True, 'message': '预开票已记录，订单继续原流程', 'invoice': _invoice_payload(invoice)}, json_dumps_params={'ensure_ascii': False})


def _action_invoice_create(request, payload):
    role_error = _require_role(request.user, ROLE_ACCOUNTING)
    if role_error:
        return role_error
    report, error = _get_report(payload)
    if error:
        return error
    if report.report_status != TestReport.Status.APPROVED:
        return JsonResponse({'ok': False, 'error': '只有终审通过的报告可以开票'}, status=400, json_dumps_params={'ensure_ascii': False})
    if report.invoices.filter(invoice_stage=Invoice.Stage.FINAL).exists():
        return JsonResponse({'ok': False, 'error': '该报告已开票'}, status=400, json_dumps_params={'ensure_ascii': False})
    with transaction.atomic():
        order = LabOrder.objects.select_for_update().get(pk=report.order_id)
        if Invoice.objects.filter(report=report, invoice_stage=Invoice.Stage.FINAL).exists():
            return JsonResponse({'ok': False, 'error': '该报告已完成最终总开票'}, status=400, json_dumps_params={'ensure_ascii': False})
        invoiced_total, remaining_amount = _invoice_amounts(order)
        amount, amount_error = _invoice_amount_from_payload(payload, default_amount=remaining_amount)
        if amount_error:
            return JsonResponse({'ok': False, 'error': amount_error}, status=400, json_dumps_params={'ensure_ascii': False})
        if amount > remaining_amount:
            return JsonResponse(
                {'ok': False, 'error': f'开票累计金额不能超过订单金额；当前剩余 {remaining_amount:.2f} 元'},
                status=400,
                json_dumps_params={'ensure_ascii': False},
            )
        pay_status, pay_error = _invoice_payment_status(payload)
        if pay_error:
            return JsonResponse({'ok': False, 'error': pay_error}, status=400, json_dumps_params={'ensure_ascii': False})
        invoice_no = (payload.get('invoice_no') or _next_invoice_no(order)).strip()
        if Invoice.objects.filter(invoice_no=invoice_no).exists():
            return JsonResponse({'ok': False, 'error': '发票号码已存在'}, status=400, json_dumps_params={'ensure_ascii': False})
        invoice = Invoice.objects.create(
            order=order,
            report=report,
            invoice_stage=Invoice.Stage.FINAL,
            invoice_no=invoice_no,
            invoice_amount=amount,
            invoice_type=payload.get('invoice_type') or '增值税专票',
            invoice_date=_parse_datetime(payload.get('invoice_date')) or timezone.now(),
            pay_status=pay_status,
            finance_user=request.user,
            order_finish_flag=Invoice.FinishFlag.FINISHED,
        )
        _event(
            order,
            request.user,
            f'最终总开票：{invoice.invoice_no}，金额 {invoice.invoice_amount:.2f} 元',
            action_code='finance_final_invoice',
            changes={
                'invoice_no': _audit_change('发票号', '', invoice.invoice_no),
                'invoice_amount': _audit_change('累计已开', invoiced_total, invoiced_total + invoice.invoice_amount),
                'remaining_amount': _audit_change('剩余可开', remaining_amount, remaining_amount - invoice.invoice_amount),
            },
        )
        order.mark_status(LabOrder.Status.INVOICED_CLOSED, request.user, f'会计最终总开票办结：{invoice.invoice_no}')
    return JsonResponse({'ok': True, 'message': '开票办结完成', 'invoice': _invoice_payload(invoice)}, json_dumps_params={'ensure_ascii': False})


def _action_invoice_pay(request, payload):
    role_error = _require_role(request.user, ROLE_ACCOUNTING)
    if role_error:
        return role_error
    invoice_no = (payload.get('invoice_no') or '').strip()
    try:
        invoice = Invoice.objects.select_related('order', 'report', 'finance_user').get(invoice_no=invoice_no)
    except Invoice.DoesNotExist:
        return JsonResponse({'ok': False, 'error': '发票不存在'}, status=404, json_dumps_params={'ensure_ascii': False})
    pay_status, pay_error = _invoice_payment_status(payload)
    if pay_error:
        return JsonResponse({'ok': False, 'error': pay_error}, status=400, json_dumps_params={'ensure_ascii': False})
    before_status = invoice.get_pay_status_display()
    invoice.pay_status = pay_status
    invoice.finance_user = request.user
    invoice.save()
    _event(
        invoice.order,
        request.user,
        f'会计更新回款状态：{invoice.invoice_no}',
        action_code='finance_payment_update',
        changes={
            'pay_status': _audit_change('回款状态', before_status, invoice.get_pay_status_display()),
        },
    )
    return JsonResponse({'ok': True, 'message': '回款状态已更新', 'invoice': _invoice_payload(invoice)}, json_dumps_params={'ensure_ascii': False})


def lims_dashboard(request):
    if not request.user.is_authenticated:
        return JsonResponse({'ok': False, 'error': '请先登录'}, status=401, json_dumps_params={'ensure_ascii': False})

    related_orders = _orders_for_user(request.user)
    active_related_orders = related_orders.exclude(
        order_status__in=[LabOrder.Status.INVOICED_CLOSED, LabOrder.Status.CANCELLED]
    )
    try:
        list_limit = int(request.GET.get('limit', 50))
    except (TypeError, ValueError):
        list_limit = 50
    list_limit = max(10, min(list_limit, 100))
    recent_orders = [
        _order_payload(order)
        for order in related_orders.order_by('-create_time')[:10]
    ]
    all_orders = [_order_payload(order) for order in related_orders.order_by('-create_time')]
    active_orders = [
        _order_payload(order)
        for order in active_related_orders.order_by('-create_time')
    ]
    report_orders = [
        _order_payload(order)
        for order in related_orders.filter(order_status=LabOrder.Status.REPORT_REVIEW).order_by('-create_time')
    ]
    running_order_ids = Experiment.objects.filter(
        order__in=related_orders,
        test_status=Experiment.Status.RUNNING,
    ).values_list('order_id', flat=True)
    running_orders = [
        _order_payload(order)
        for order in related_orders.filter(id__in=running_order_ids).order_by('-create_time')
    ]
    change_order_ids = ChangeRequest.objects.filter(
        order__in=related_orders,
    ).exclude(change_status=ChangeRequest.Status.APPLIED).values_list('order_id', flat=True)
    change_orders = [
        _order_payload(order)
        for order in related_orders.filter(id__in=change_order_ids).order_by('-create_time')
    ]
    can_view_finance = _can_view_finance(request.user)
    finance_order_ids = []
    preinvoice_candidate_orders = []
    if can_view_finance:
        preinvoice_scope = related_orders.filter(order_status__in=[
            LabOrder.Status.SCHEDULING,
            LabOrder.Status.TESTING,
            LabOrder.Status.RESULT_PENDING,
            LabOrder.Status.TEST_FINISHED,
            LabOrder.Status.REPORT_REVIEW,
        ]).prefetch_related('reviews', 'reports', 'invoices', 'experiments').order_by('-create_time')
        for order in preinvoice_scope:
            if _preinvoice_stage(order):
                _, remaining_amount = _invoice_amounts(order)
                if remaining_amount > Decimal('0.01'):
                    preinvoice_candidate_orders.append(order)
        finance_order_ids = list(
            TestReport.objects.filter(
                order__in=related_orders,
                report_status=TestReport.Status.APPROVED,
            ).exclude(
                invoices__invoice_stage=Invoice.Stage.FINAL,
            ).values_list('order_id', flat=True)
        ) + list(
            Invoice.objects.filter(order__in=related_orders).values_list('order_id', flat=True)
        ) + [order.id for order in preinvoice_candidate_orders]
    finance_orders = [
        _order_payload(order)
        for order in related_orders.filter(id__in=finance_order_ids).distinct().order_by('-create_time')
    ]
    pending_reports = _pending_reports_for_user(request.user, related_orders)
    outsource_orders = [
        _order_payload(order)
        for order in related_orders.filter(
            Q(execution_mode=LabOrder.ExecutionMode.OUTSOURCE)
            | Q(schedules__test_type=SchedulePlan.TestType.OUTSOURCE)
        ).distinct().order_by('-create_time')
    ]
    schedules = SchedulePlan.objects.select_related('order', 'lab_manager', 'quality_user', 'device').prefetch_related(
        Prefetch('sample_photos', queryset=SamplePhoto.objects.order_by('create_time', 'id'), to_attr='ordered_sample_photos'),
        Prefetch(
            'samples',
            queryset=Sample.objects.select_related('quality_user', 'outbound_by').order_by('id'),
            to_attr='ordered_samples',
        ),
        Prefetch(
            'experiments',
            queryset=Experiment.objects.select_related('test_operator').order_by('-create_time'),
            to_attr='ordered_experiments',
        ),
    ).filter(
        order__in=related_orders
    ).order_by('-plan_start_time', '-create_time')
    samples = Sample.objects.select_related('order', 'schedule', 'quality_user', 'outbound_by').prefetch_related(
        'schedule__sample_photos'
    ).filter(
        order__in=related_orders
    ).order_by('-actual_arrive_time', '-create_time')
    changes = ChangeRequest.objects.select_related('order', 'schedule', 'change_user').filter(
        order__in=related_orders
    ).order_by('-change_time', '-create_time')
    dashboard_roles = set(_roles(request.user))
    if not _is_chairman(request.user) and (ROLE_SUZHOU_LAB in dashboard_roles or ROLE_JIANGYIN_LAB in dashboard_roles):
        schedules = schedules.filter(lab_manager=request.user)
        samples = samples.filter(schedule__lab_manager=request.user)
        changes = changes.filter(schedule__lab_manager=request.user)
    elif not _is_chairman(request.user) and ROLE_LAB_OPERATOR in dashboard_roles:
        lab_type = _user_lab_type(request.user)
        schedule_scope = _lab_schedule_query(lab_type) if lab_type else Q(pk__in=[])
        schedules = schedules.filter(schedule_scope)
        samples = samples.filter(schedule__in=SchedulePlan.objects.filter(schedule_scope))
        changes = changes.filter(schedule__in=SchedulePlan.objects.filter(schedule_scope))
    reviews = BusinessReview.objects.select_related('order', 'biz_review_user', 'tech_review_user').filter(
        order__in=related_orders
    ).order_by('-review_time', '-create_time')
    workflow_events = WorkflowEvent.objects.select_related('order', 'actor', 'schedule').filter(
        order__in=related_orders
    ).order_by('-create_time')[:120]
    test_standards = TestStandard.objects.filter(is_active=True).order_by('industry', 'standard_code')
    pending_invoice_reports = TestReport.objects.none()
    invoices = Invoice.objects.none()
    if can_view_finance:
        pending_invoice_reports = TestReport.objects.select_related('order').prefetch_related(
            'order__invoices', 'order__experiments',
        ).filter(
            order__in=related_orders,
            report_status=TestReport.Status.APPROVED,
        ).exclude(
            invoices__invoice_stage=Invoice.Stage.FINAL,
        ).order_by('-create_time')
        invoices = Invoice.objects.select_related('order', 'report', 'finance_user').prefetch_related(
            'order__invoices', 'order__experiments',
        ).filter(
            order__in=related_orders,
        ).order_by('-invoice_date', '-create_time')

    data = {
        'company': '苏州环测检测技术有限公司',
        'system': '实验室管理（LIMS）系统',
        'metrics': {
            'orders': related_orders.count(),
            'active_orders': active_related_orders.count(),
            'running_experiments': Experiment.objects.filter(order__in=related_orders, test_status=Experiment.Status.RUNNING).count(),
            'pending_reports': TestReport.objects.filter(order__in=related_orders).exclude(report_status=TestReport.Status.APPROVED).count(),
            'change_requests': ChangeRequest.objects.filter(order__in=related_orders).exclude(
                change_status=ChangeRequest.Status.APPLIED
            ).count(),
        },
        'payload_limits': {
            'list_limit': list_limit,
            'workflow_events': 120,
            'note': 'Dashboard returns limited recent rows; use dedicated paginated APIs for full history in future iterations.',
        },
        'status_counts': {
            item['order_status']: item['count']
            for item in related_orders.values('order_status').annotate(count=Count('id'))
        },
        'mode_counts': {
            item['execution_mode']: item['count']
            for item in related_orders.values('execution_mode').annotate(count=Count('id'))
        },
        'recent_orders': recent_orders,
        'order_groups': {
            'orders': all_orders[:list_limit],
            'active_orders': active_orders[:list_limit],
            'running_experiments': running_orders[:list_limit],
            'pending_reports': report_orders[:list_limit],
            'change_requests': change_orders[:list_limit],
            'finance_orders': finance_orders[:list_limit],
        },
        'labs': {
            'suzhou': _lab_payload(SchedulePlan.TestType.SUZHOU, '苏州实验室', related_orders, request.user),
            'jiangyin': _lab_payload(SchedulePlan.TestType.JIANGYIN, '江阴实验室', related_orders, request.user),
        },
        'outsource_orders': outsource_orders[:list_limit],
        'schedules': [_schedule_payload(schedule) for schedule in _limit_queryset(schedules, list_limit)],
        'samples': [_sample_payload(sample) for sample in _limit_queryset(samples, list_limit)],
        'changes': [_change_payload(change) for change in _limit_queryset(changes, list_limit)],
        'reviews': [_review_payload(review) for review in _limit_queryset(reviews, list_limit)],
        'workflow_events': [_workflow_payload(event) for event in workflow_events],
        'test_standards': [_standard_payload(standard) for standard in test_standards],
        'routing_options': {
            'suzhou_managers': _role_user_options(ROLE_SUZHOU_LAB),
            'jiangyin_managers': _role_user_options(ROLE_JIANGYIN_LAB),
        } if _has_any_role(request.user, ROLE_TECH) else {
            'suzhou_managers': [],
            'jiangyin_managers': [],
        },
        'pending_reports': [_report_payload(report) for report in _limit_queryset(pending_reports.order_by('-create_time'), list_limit)],
        'finance': {
            'preinvoice_candidates': [
                _pending_preinvoice_payload(order)
                for order in preinvoice_candidate_orders[:list_limit]
            ],
            'pending_invoices': [_pending_final_invoice_payload(report) for report in _limit_queryset(pending_invoice_reports, list_limit)],
            'issued_invoices': [_invoice_payload(invoice) for invoice in _limit_queryset(invoices, list_limit)],
        },
        'roles': _roles(request.user),
    }
    return JsonResponse(data, json_dumps_params={'ensure_ascii': False})
