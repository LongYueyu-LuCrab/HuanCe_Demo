from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase
from django.urls import reverse
from django.utils import timezone
from openpyxl import load_workbook
from io import BytesIO

from .models import (
    BusinessReview,
    ChangeRequest,
    Experiment,
    Invoice,
    LabDevice,
    LabOrder,
    LabStaffProfile,
    OrderDocument,
    ReportAudit,
    Sample,
    SchedulePlan,
    TestReport,
    WorkflowEvent,
)


class LimsDashboardTests(TestCase):
    def setUp(self):
        self.user = get_user_model().objects.create_user(
            username='tester',
            password='password123',
        )
        sales_group = Group.objects.create(name='销售')
        self.user.groups.add(sales_group)
        self.order = LabOrder.objects.create(
            order_no='TEST-001',
            customer_name='苏州环测检测技术有限公司',
            customer_contact='王五',
            customer_phone='13800000000',
            project_name='测试订单',
            test_demand='测试需求',
            test_method='振动扫频后进行定频耐久试验',
            test_standard='GB/T 2423.10-2019',
            sale_user=self.user,
            order_status=LabOrder.Status.TESTING,
        )

    def test_dashboard_api_returns_lims_summary(self):
        self.client.force_login(self.user)
        response = self.client.get(reverse('lims_dashboard'))

        self.assertEqual(response.status_code, 200)
        data = response.json()
        self.assertEqual(data['company'], '苏州环测检测技术有限公司')
        self.assertEqual(data['metrics']['orders'], 1)
        self.assertEqual(data['metrics']['active_orders'], 1)
        self.assertEqual(len(data['recent_orders']), 1)
        self.assertEqual(data['recent_orders'][0]['order_no'], 'TEST-001')

    def test_order_detail_returns_complete_visible_context(self):
        self.order.industry_category = LabOrder.IndustryCategory.MILITARY
        self.order.autonomous_execution = True
        self.order.outsourced_execution = True
        self.order.remark = '加急；完整信息测试'
        self.order.save()
        self.client.force_login(self.user)

        response = self.client.get(
            reverse('order_detail', kwargs={'order_no': self.order.order_no})
        )

        self.assertEqual(response.status_code, 200)
        order = response.json()['order']
        self.assertEqual(order['industry_label'], '军工')
        self.assertEqual(order['execution_attributes'], ['自主', '委外'])
        self.assertEqual(order['contact'], self.order.customer_contact)
        self.assertEqual(order['test_method'], '振动扫频后进行定频耐久试验')
        self.assertEqual(order['test_standard'], 'GB/T 2423.10-2019')
        self.assertEqual(order['remark'], '加急；完整信息测试')
        self.assertIn('documents', order)

    def test_order_detail_respects_sales_row_permission(self):
        other_user = get_user_model().objects.create_user(
            username='other-sales',
            password='password123',
        )
        other_user.groups.add(Group.objects.get(name='销售'))
        self.client.force_login(other_user)

        response = self.client.get(
            reverse('order_detail', kwargs={'order_no': self.order.order_no})
        )

        self.assertEqual(response.status_code, 404)

    def test_technical_reviewer_sees_methods_standards_documents_and_execution_attributes(self):
        technical_user = get_user_model().objects.create_user(
            username='technical-reviewer',
            password='password123',
        )
        technical_user.groups.add(Group.objects.create(name='技术'))
        self.order.order_status = LabOrder.Status.PENDING_REVIEW
        self.order.autonomous_execution = True
        self.order.outsourced_execution = True
        self.order.save()
        OrderDocument.objects.create(
            order=self.order,
            document_type=OrderDocument.DocumentType.ATTACHMENT,
            file=SimpleUploadedFile('技术要求.pdf', b'%PDF-1.4 technical'),
            original_name='技术要求.pdf',
            file_size=18,
            uploaded_by=self.user,
        )
        self.client.force_login(technical_user)

        response = self.client.get(
            reverse('order_detail', kwargs={'order_no': self.order.order_no})
        )

        self.assertEqual(response.status_code, 200)
        order = response.json()['order']
        self.assertEqual(order['test_method'], self.order.test_method)
        self.assertEqual(order['test_standard'], self.order.test_standard)
        self.assertEqual(order['execution_attributes'], ['自主', '委外'])
        self.assertEqual(order['documents'][0]['name'], '技术要求.pdf')

    def test_mark_status_writes_workflow_event(self):
        self.order.mark_status(LabOrder.Status.REPORT_REVIEW, note='测试流转')

        self.order.refresh_from_db()
        self.assertEqual(self.order.order_status, LabOrder.Status.REPORT_REVIEW)
        self.assertTrue(
            WorkflowEvent.objects.filter(
                order=self.order,
                from_status=str(LabOrder.Status.TESTING),
                to_status=str(LabOrder.Status.REPORT_REVIEW),
            ).exists()
        )

    def test_sales_can_create_order(self):
        self.client.force_login(self.user)

        response = self.client.post(
            reverse('create_order'),
            data={
                'customer_name': '新客户',
                'contact_name': '王五',
                'phone': '13800000000',
                'project_name': '新项目',
                'test_requirements': '完成可靠性检测。',
                'test_method': '按规定载荷完成振动耐久试验。',
                'test_standard': 'GB/T 2423.10-2019',
                'expected_sample_arrival': '2026-07-01',
                'expected_delivery_date': '2026-07-15',
                'quoted_amount': '12000.50',
                'is_urgent': True,
                'industry_category': 'automotive',
                'execution_attributes': ['autonomous', 'outsource'],
            },
            content_type='application/json',
        )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertTrue(payload['ok'])
        created = LabOrder.objects.get(order_no=payload['order']['order_no'])
        self.assertEqual(created.sale_user, self.user)
        self.assertEqual(created.order_status, LabOrder.Status.PENDING_REVIEW)
        self.assertEqual(created.customer_name, '新客户')
        self.assertEqual(created.industry_category, LabOrder.IndustryCategory.AUTOMOTIVE)
        self.assertEqual(created.test_method, '按规定载荷完成振动耐久试验。')
        self.assertEqual(created.test_standard, 'GB/T 2423.10-2019')
        self.assertTrue(created.autonomous_execution)
        self.assertTrue(created.outsourced_execution)

    def test_sales_can_upload_contract_and_attachment(self):
        self.client.force_login(self.user)
        contract = SimpleUploadedFile('合同.pdf', b'%PDF-1.4 demo', content_type='application/pdf')
        attachment = SimpleUploadedFile('需求说明.docx', b'word-demo', content_type='application/vnd.openxmlformats-officedocument.wordprocessingml.document')

        response = self.client.post(
            reverse('create_order'),
            data={
                'customer_name': '附件测试客户',
                'project_name': '文件上传测试',
                'test_requirements': '验证合同与附件上传。',
                'expected_sample_arrival': '2026-07-01',
                'industry_category': 'military',
                'execution_attributes': ['autonomous'],
                'contract_files': contract,
                'attachment_files': attachment,
            },
        )

        self.assertEqual(response.status_code, 200)
        order = LabOrder.objects.get(order_no=response.json()['order']['order_no'])
        self.assertEqual(order.documents.count(), 2)
        contract_record = order.documents.get(document_type=OrderDocument.DocumentType.CONTRACT)
        self.assertEqual(contract_record.original_name, '合同.pdf')

        download = self.client.get(
            reverse('download_order_document', kwargs={'document_id': contract_record.id})
        )
        self.assertEqual(download.status_code, 200)
        self.assertEqual(download['Content-Disposition'], "attachment; filename*=utf-8''%E5%90%88%E5%90%8C.pdf")

    def test_sales_order_requires_expected_sample_arrival(self):
        self.client.force_login(self.user)
        response = self.client.post(
            reverse('create_order'),
            data={
                'customer_name': '缺少到样时间客户',
                'project_name': '必填校验',
                'test_requirements': '验证预计到样时间必填。',
                'industry_category': 'other',
                'execution_attributes': ['autonomous'],
            },
            content_type='application/json',
        )
        self.assertEqual(response.status_code, 400)
        self.assertIn('预计样品到达时间必填', response.json()['error'])

    def test_order_upload_rejects_unsupported_file_type(self):
        self.client.force_login(self.user)
        unsafe_file = SimpleUploadedFile('程序.exe', b'not-allowed')

        response = self.client.post(
            reverse('create_order'),
            data={
                'customer_name': '非法文件客户',
                'project_name': '非法文件测试',
                'test_requirements': '验证上传格式限制。',
                'expected_sample_arrival': '2026-07-01',
                'industry_category': 'other',
                'execution_attributes': ['outsource'],
                'attachment_files': unsafe_file,
            },
        )

        self.assertEqual(response.status_code, 400)
        self.assertIn('格式不支持', response.json()['error'])


class LimsFullRoleWorkflowTests(TestCase):
    roles = {
        'sales': '销售',
        'business': '商务',
        'tech': '技术',
        'quality': '质量部',
        'suzhou_lab': '苏州实验室',
        'jiangyin_lab': '江阴实验室',
        'general_manager': '总经理',
        'accountant': '会计',
    }

    def setUp(self):
        user_model = get_user_model()
        self.users = {}
        for username, role_name in self.roles.items():
            group = Group.objects.create(name=role_name)
            user = user_model.objects.create_user(
                username=username,
                password='password123',
                first_name=role_name,
            )
            user.groups.add(group)
            self.users[username] = user

    def test_all_roles_can_complete_one_lims_order_flow(self):
        now = timezone.now()

        order = LabOrder.objects.create(
            order_no='FLOW-001',
            customer_name='流程测试客户',
            customer_contact='赵六',
            customer_phone='13900000000',
            project_name='全角色流程测试',
            test_demand='覆盖销售、商务、技术、质量、双实验室、总经理、会计的完整流程。',
            total_quote='50000.00',
            expect_sample_arrive=now + timezone.timedelta(days=1),
            sale_user=self.users['sales'],
            create_by=self.users['sales'].username,
            update_by=self.users['sales'].username,
            order_status=LabOrder.Status.PENDING_REVIEW,
            workflow_version=LabOrder.WorkflowVersion.LEGACY_QUALITY,
        )
        WorkflowEvent.objects.create(
            order=order,
            actor=self.users['sales'],
            event_type=WorkflowEvent.EventType.STATUS,
            to_status=str(LabOrder.Status.PENDING_REVIEW),
            note='销售下单',
        )

        review = BusinessReview.objects.create(
            order=order,
            biz_review_user=self.users['business'],
            tech_review_user=self.users['tech'],
            biz_quote_detail='商务报价通过',
            tech_feasible=True,
            review_result=True,
            review_time=now,
        )
        order.mark_status(LabOrder.Status.SCHEDULING, self.users['business'], '商务技术评审通过')

        suzhou_schedule = SchedulePlan.objects.create(
            order=order,
            test_type=SchedulePlan.TestType.SUZHOU,
            lab_manager=self.users['suzhou_lab'],
            plan_start_time=now + timezone.timedelta(days=2),
            plan_end_time=now + timezone.timedelta(days=5),
            schedule_status=SchedulePlan.Status.RUNNING,
            quality_user=self.users['quality'],
            remark='苏州实验室排期',
        )
        jiangyin_schedule = SchedulePlan.objects.create(
            order=order,
            test_type=SchedulePlan.TestType.JIANGYIN,
            lab_manager=self.users['jiangyin_lab'],
            plan_start_time=now + timezone.timedelta(days=3),
            plan_end_time=now + timezone.timedelta(days=6),
            schedule_status=SchedulePlan.Status.RUNNING,
            quality_user=self.users['quality'],
            remark='江阴实验室排期',
        )

        change = ChangeRequest.objects.create(
            order=order,
            schedule=suzhou_schedule,
            change_scene=ChangeRequest.Scene.BEFORE_SAMPLE,
            old_test_demand=order.test_demand,
            new_test_demand=f'{order.test_demand} 增加温湿度循环。',
            change_content='样品到货前客户调整检测条件',
            change_user=self.users['sales'],
            change_status=ChangeRequest.Status.APPLIED,
        )
        suzhou_schedule.plan_end_time = suzhou_schedule.plan_end_time + timezone.timedelta(days=1)
        suzhou_schedule.save(update_fields=['plan_end_time', 'update_time'])

        suzhou_sample = Sample.objects.create(
            order=order,
            schedule=suzhou_schedule,
            sample_no='SAMPLE-SZ-001',
            sample_name='苏州试验样品',
            sample_spec='A 型',
            sample_count=2,
            storage_condition='常温',
            actual_arrive_time=now,
            sample_status=Sample.Status.TESTING,
            quality_user=self.users['quality'],
        )
        jiangyin_sample = Sample.objects.create(
            order=order,
            schedule=jiangyin_schedule,
            sample_no='SAMPLE-JY-001',
            sample_name='江阴试验样品',
            sample_spec='B 型',
            sample_count=1,
            storage_condition='避光',
            actual_arrive_time=now,
            sample_status=Sample.Status.TESTING,
            quality_user=self.users['quality'],
        )
        order.mark_status(LabOrder.Status.TESTING, self.users['quality'], '样品登记完成，开始试验')

        suzhou_test = Experiment.objects.create(
            order=order,
            schedule=suzhou_schedule,
            sample=suzhou_sample,
            test_item_list='高低温循环试验',
            test_standard='GB/T 2423',
            test_start_time=now + timezone.timedelta(days=2),
            test_end_time=now + timezone.timedelta(days=5),
            test_operator=self.users['suzhou_lab'],
            test_raw_data='苏州实验室原始数据',
            test_conclusion_temp='苏州项目合格',
            test_status=Experiment.Status.FINISHED,
            test_type=SchedulePlan.TestType.SUZHOU,
        )
        jiangyin_test = Experiment.objects.create(
            order=order,
            schedule=jiangyin_schedule,
            sample=jiangyin_sample,
            test_item_list='振动耐久试验',
            test_standard='GB/T 2423',
            test_start_time=now + timezone.timedelta(days=3),
            test_end_time=now + timezone.timedelta(days=6),
            test_operator=self.users['jiangyin_lab'],
            test_raw_data='江阴实验室原始数据',
            test_conclusion_temp='江阴项目合格',
            test_status=Experiment.Status.FINISHED,
            test_type=SchedulePlan.TestType.JIANGYIN,
        )
        self.assertNotEqual(suzhou_test.test_operator, jiangyin_test.test_operator)

        report = TestReport.objects.create(
            order=order,
            test_record=suzhou_test,
            report_no='RPT-FLOW-001',
            report_file_url='/media/reports/RPT-FLOW-001.pdf',
            final_conclusion='全部检测项目符合要求。',
            report_status=TestReport.Status.SALES_REVIEW,
            create_quality_user=self.users['quality'],
        )
        order.mark_status(LabOrder.Status.REPORT_REVIEW, self.users['quality'], '质量部出具报告')

        sales_audit = ReportAudit.objects.create(
            report=report,
            audit_level=ReportAudit.Level.SALES,
            audit_user=self.users['sales'],
            audit_result=ReportAudit.Result.APPROVED,
            audit_opinion='销售初审通过',
            audit_time=now,
        )
        report.report_status = TestReport.Status.GM_REVIEW
        report.save(update_fields=['report_status', 'update_time'])

        gm_audit = ReportAudit.objects.create(
            report=report,
            audit_level=ReportAudit.Level.GENERAL_MANAGER,
            audit_user=self.users['general_manager'],
            audit_result=ReportAudit.Result.APPROVED,
            audit_opinion='总经理终审通过',
            audit_time=now,
        )
        report.report_status = TestReport.Status.APPROVED
        report.save(update_fields=['report_status', 'update_time'])

        invoice = Invoice.objects.create(
            order=order,
            report=report,
            invoice_no='INV-FLOW-001',
            invoice_amount=order.total_quote,
            invoice_type='增值税专票',
            invoice_date=now,
            pay_status=Invoice.PayStatus.PAID,
            finance_user=self.users['accountant'],
            order_finish_flag=Invoice.FinishFlag.FINISHED,
        )
        order.mark_status(LabOrder.Status.INVOICED_CLOSED, self.users['accountant'], '会计开票办结')

        order.refresh_from_db()
        report.refresh_from_db()

        self.assertTrue(review.review_result)
        self.assertEqual(change.change_status, ChangeRequest.Status.APPLIED)
        self.assertEqual(order.schedules.count(), 2)
        self.assertEqual(order.samples.count(), 2)
        self.assertEqual(order.experiments.count(), 2)
        self.assertEqual(report.audits.count(), 2)
        self.assertEqual(sales_audit.audit_user, self.users['sales'])
        self.assertEqual(gm_audit.audit_user, self.users['general_manager'])
        self.assertEqual(invoice.finance_user, self.users['accountant'])
        self.assertEqual(invoice.order_finish_flag, Invoice.FinishFlag.FINISHED)
        self.assertEqual(report.report_status, TestReport.Status.APPROVED)
        self.assertEqual(order.order_status, LabOrder.Status.INVOICED_CLOSED)
        self.assertEqual(
            set(Group.objects.values_list('name', flat=True)),
            set(self.roles.values()) | {'实验操作员'},
        )


class LimsV2DirectLabWorkflowTests(TestCase):
    def setUp(self):
        user_model = get_user_model()
        self.users = {}
        for username, role in {
            'sales_v2': '销售',
            'business_v2': '商务',
            'tech_v2': '技术',
            'quality_v1': '质量部',
            'suzhou_v2': '苏州实验室',
            'jiangyin_v2': '江阴实验室',
        }.items():
            group, _ = Group.objects.get_or_create(name=role)
            user = user_model.objects.create_user(username=username, password='password123', first_name=role)
            user.groups.add(group)
            self.users[username] = user
        self.order = LabOrder.objects.create(
            order_no='FLOW-V2-001',
            customer_name='V2 流程客户',
            project_name='多路径直达实验室验证',
            test_demand='苏州环境试验、江阴振动试验及委外盐雾试验',
            test_standard='GB/T 2423',
            total_quote='68000.00',
            autonomous_execution=True,
            outsourced_execution=True,
            sale_user=self.users['sales_v2'],
            workflow_version=LabOrder.WorkflowVersion.LAB_DIRECT,
        )
        self.suzhou_device = LabDevice.objects.create(
            device_code='TEST-SZ-001', device_name='苏州测试台', lab_type=LabDevice.LabType.SUZHOU,
        )
        self.jiangyin_device = LabDevice.objects.create(
            device_code='TEST-JY-001', device_name='江阴测试台', lab_type=LabDevice.LabType.JIANGYIN,
        )

    def action(self, user_key, action, **payload):
        self.client.force_login(self.users[user_key])
        data = {'action': action, 'order_no': self.order.order_no, **payload}
        if 'sample_photos' in payload:
            data = {key: value for key, value in data.items() if value is not None}
            return self.client.post(reverse('lims_action'), data=data)
        return self.client.post(
            reverse('lims_action'),
            data=data,
            content_type='application/json',
        )

    def test_v2_routes_directly_to_labs_and_lead_manager_issues_report(self):
        business = self.action('business_v2', 'review_pass', biz_quote_detail='商务评审通过')
        self.assertEqual(business.status_code, 200)

        technical = self.action(
            'tech_v2',
            'review_pass',
            biz_quote_detail='技术可行，多路径执行',
            execution_routes=['suzhou', 'jiangyin', 'outsource'],
            suzhou_manager_id=self.users['suzhou_v2'].id,
            jiangyin_manager_id=self.users['jiangyin_v2'].id,
            outsource_owner_id=self.users['suzhou_v2'].id,
            lead_lab_manager_id=self.users['suzhou_v2'].id,
            suzhou_task='高低温循环',
            jiangyin_task='振动耐久',
            outsource_task='盐雾腐蚀',
        )
        self.assertEqual(technical.status_code, 200)
        self.order.refresh_from_db()
        self.assertEqual(self.order.order_status, LabOrder.Status.SCHEDULING)
        self.assertEqual(self.order.lead_lab_manager, self.users['suzhou_v2'])
        self.assertEqual(self.order.schedules.count(), 3)
        self.assertEqual(self.order.schedules.filter(assigned_by=self.users['tech_v2']).count(), 3)

        suzhou_schedule = self.order.schedules.get(test_type=SchedulePlan.TestType.SUZHOU)
        jiangyin_schedule = self.order.schedules.get(test_type=SchedulePlan.TestType.JIANGYIN)
        outsource_schedule = self.order.schedules.get(test_type=SchedulePlan.TestType.OUTSOURCE)

        changed = self.action(
            'sales_v2', 'create_change', change_scene=1,
            new_test_demand='三条路径均增加复测要求', change_content='客户统一调整试验要求',
        )
        self.assertEqual(changed.status_code, 200)
        self.assertEqual(self.order.change_requests.filter(change_status=ChangeRequest.Status.PENDING).count(), 3)
        for user_key, schedule in (
            ('suzhou_v2', suzhou_schedule),
            ('jiangyin_v2', jiangyin_schedule),
            ('suzhou_v2', outsource_schedule),
        ):
            self.assertEqual(
                self.action(
                    user_key, 'process_change', schedule_id=schedule.id,
                    plan_start_time='2026-08-29', plan_end_time='2026-08-31',
                    device_id=(
                        self.suzhou_device.id if schedule.test_type == SchedulePlan.TestType.SUZHOU
                        else self.jiangyin_device.id if schedule.test_type == SchedulePlan.TestType.JIANGYIN
                        else None
                    ),
                ).status_code,
                200,
            )

        quality_denied = self.action(
            'quality_v1', 'schedule_assign', schedule_id=suzhou_schedule.id,
            plan_start_time='2026-09-01', plan_end_time='2026-09-05',
        )
        self.assertEqual(quality_denied.status_code, 403)

        confirmed = self.action('sales_v2', 'sales_confirm', note='销售确认无变更')
        self.assertEqual(confirmed.status_code, 200)

        for user_key, schedule in (
            ('suzhou_v2', suzhou_schedule),
            ('jiangyin_v2', jiangyin_schedule),
            ('suzhou_v2', outsource_schedule),
        ):
            schedule_response = self.action(
                user_key, 'schedule_assign', schedule_id=schedule.id,
                plan_start_time='2026-09-01', plan_end_time='2026-09-05',
                device_id=(
                    self.suzhou_device.id if schedule.test_type == SchedulePlan.TestType.SUZHOU
                    else self.jiangyin_device.id if schedule.test_type == SchedulePlan.TestType.JIANGYIN
                    else None
                ),
                outsource_factory='委外测试机构' if schedule.test_type == SchedulePlan.TestType.OUTSOURCE else '',
                sample_arrived='true',
                sample_photos=SimpleUploadedFile(
                    f'{schedule.id}.jpg', b'fake-jpeg-content', content_type='image/jpeg'
                ),
            )
            self.assertEqual(schedule_response.status_code, 200)
            schedule.refresh_from_db()
            self.assertTrue(schedule.sample_arrived)
            self.assertEqual(schedule.sample_photos.count(), 1)

        for user_key, schedule in (('suzhou_v2', suzhou_schedule), ('jiangyin_v2', jiangyin_schedule)):
            self.assertEqual(
                self.action(user_key, 'start_test', schedule_id=schedule.id, test_standard='此值应被后端忽略').status_code,
                200,
            )
            self.assertEqual(schedule.experiments.get().test_standard, 'GB/T 2423')
            self.assertEqual(
                self.action(user_key, 'submit_test', test_raw_data='原始数据完整', test_conclusion_temp='合格').status_code,
                200,
            )

        self.assertEqual(
            self.action(
                'suzhou_v2', 'outsource_result', schedule_id=outsource_schedule.id,
                test_raw_data='委外报告数据', test_conclusion_temp='合格',
            ).status_code,
            200,
        )
        denied_report = self.action('jiangyin_v2', 'issue_report', final_conclusion='全部合格')
        self.assertEqual(denied_report.status_code, 403)
        issued_report = self.action('suzhou_v2', 'issue_report', final_conclusion='三条路径全部完成，结论合格')
        self.assertEqual(issued_report.status_code, 200)
        self.order.refresh_from_db()
        self.assertEqual(self.order.order_status, LabOrder.Status.REPORT_REVIEW)
        self.assertEqual(self.order.reports.get().create_quality_user, self.users['suzhou_v2'])


class LabDeviceSchedulingTests(TestCase):
    def setUp(self):
        user_model = get_user_model()
        group = Group.objects.create(name='苏州实验室')
        self.manager = user_model.objects.create_user(username='device_manager', password='password123')
        self.manager.groups.add(group)
        self.device = LabDevice.objects.create(
            device_code='DEVICE-SZ-001', device_name='冲突验证台', lab_type=LabDevice.LabType.SUZHOU,
        )
        self.order = LabOrder.objects.create(
            order_no='DEVICE-ORDER-001', customer_name='设备排期客户', project_name='设备排期验证',
            test_demand='振动耐久试验', test_standard='GB/T 2423.10', total_quote='10000.00',
            order_status=LabOrder.Status.SCHEDULING, workflow_version=LabOrder.WorkflowVersion.LAB_DIRECT,
            lead_lab_manager=self.manager, sales_confirmed_at=timezone.now(),
        )
        self.schedule = SchedulePlan.objects.create(
            order=self.order, test_type=SchedulePlan.TestType.SUZHOU, lab_manager=self.manager,
            quality_user=self.manager, remark='振动耐久试验',
        )

    def action(self, action, **payload):
        self.client.force_login(self.manager)
        return self.client.post(
            reverse('lims_action'),
            data={'action': action, 'order_no': self.order.order_no, 'schedule_id': self.schedule.id, **payload},
            content_type='application/json',
        )

    def test_device_availability_and_conflict_are_enforced(self):
        self.client.force_login(self.manager)
        available = self.client.get(reverse('lab_device_availability'), {
            'schedule_id': self.schedule.id, 'start_date': '2026-09-01', 'end_date': '2026-09-05',
        })
        self.assertEqual(available.status_code, 200)
        selected = next(item for item in available.json()['devices'] if item['id'] == self.device.id)
        self.assertTrue(selected['available'])

        first = self.action(
            'schedule_assign', device_id=self.device.id,
            plan_start_time='2026-09-01', plan_end_time='2026-09-05',
        )
        self.assertEqual(first.status_code, 200)

        other_order = LabOrder.objects.create(
            order_no='DEVICE-ORDER-002', customer_name='冲突客户', project_name='冲突订单',
            test_demand='同台试验', total_quote='8000.00', order_status=LabOrder.Status.SCHEDULING,
            workflow_version=LabOrder.WorkflowVersion.LAB_DIRECT, lead_lab_manager=self.manager,
        )
        other_schedule = SchedulePlan.objects.create(
            order=other_order, test_type=SchedulePlan.TestType.SUZHOU, lab_manager=self.manager,
            quality_user=self.manager, remark='同台试验',
        )
        conflict = self.client.post(
            reverse('lims_action'),
            data={
                'action': 'schedule_assign', 'order_no': other_order.order_no, 'schedule_id': other_schedule.id,
                'device_id': self.device.id, 'plan_start_time': '2026-09-04', 'plan_end_time': '2026-09-08',
            },
            content_type='application/json',
        )
        self.assertEqual(conflict.status_code, 400)
        self.assertIn('排期冲突', conflict.json()['error'])

    def test_device_crud_and_historical_delete_protection(self):
        self.client.force_login(self.manager)
        created = self.client.post(
            reverse('lab_devices'),
            data={'device_code': 'DEVICE-SZ-NEW', 'device_name': '新设备', 'model_spec': 'TEST'},
            content_type='application/json',
        )
        self.assertEqual(created.status_code, 201)
        device_id = created.json()['device']['id']
        updated = self.client.patch(
            reverse('lab_device_detail', args=[device_id]),
            data={'device_name': '新设备（已维护）', 'device_status': LabDevice.Status.MAINTENANCE},
            content_type='application/json',
        )
        self.assertEqual(updated.status_code, 200)
        self.assertEqual(updated.json()['device']['status'], '维修中')
        self.assertEqual(self.client.delete(reverse('lab_device_detail', args=[device_id])).status_code, 200)

        self.schedule.device = self.device
        self.schedule.save(update_fields=['device', 'update_time'])
        protected = self.client.delete(reverse('lab_device_detail', args=[self.device.id]))
        self.assertEqual(protected.status_code, 400)
        self.assertIn('不能删除', protected.json()['error'])


class LaboratoryOperatorTests(TestCase):
    def setUp(self):
        user_model = get_user_model()
        manager_group, _ = Group.objects.get_or_create(name='苏州实验室')
        operator_group, _ = Group.objects.get_or_create(name='实验操作员')
        self.manager = user_model.objects.create_user(username='operator_manager', password='password123')
        self.manager.groups.add(manager_group)
        LabStaffProfile.objects.update_or_create(
            user=self.manager,
            defaults={'lab_type': LabDevice.LabType.SUZHOU, 'position': LabStaffProfile.Position.MANAGER},
        )
        self.operator = user_model.objects.create_user(username='lab_operator', password='password123')
        self.operator.groups.add(operator_group)
        LabStaffProfile.objects.create(
            user=self.operator,
            lab_type=LabDevice.LabType.SUZHOU,
            position=LabStaffProfile.Position.OPERATOR,
        )
        self.device = LabDevice.objects.create(
            device_code='OP-SZ-001', device_name='操作员测试台', lab_type=LabDevice.LabType.SUZHOU,
        )
        self.order = LabOrder.objects.create(
            order_no='OPERATOR-FLOW-001', customer_name='操作员流程客户', project_name='操作员权限验证',
            test_demand='振动耐久试验', test_standard='GB/T 2423.10', total_quote='12000.00',
            order_status=LabOrder.Status.SCHEDULING, workflow_version=LabOrder.WorkflowVersion.LAB_DIRECT,
            lead_lab_manager=self.manager, sales_confirmed_at=timezone.now(),
        )
        self.schedule = SchedulePlan.objects.create(
            order=self.order, test_type=SchedulePlan.TestType.SUZHOU, lab_manager=self.manager,
            quality_user=self.manager, remark='振动耐久试验',
        )

    def action(self, action, **payload):
        self.client.force_login(self.operator)
        data = {'action': action, 'order_no': self.order.order_no, 'schedule_id': self.schedule.id, **payload}
        if 'sample_photos' in payload:
            return self.client.post(reverse('lims_action'), data=data)
        return self.client.post(
            reverse('lims_action'),
            data=data,
            content_type='application/json',
        )

    def test_operator_runs_assigned_lab_flow_and_writes_structured_history(self):
        self.assertEqual(self.action(
            'schedule_assign', device_id=self.device.id,
            plan_start_time='2026-10-01', plan_end_time='2026-10-03',
            sample_arrived='true',
            sample_photos=SimpleUploadedFile('operator-sample.png', b'fake-png-content', content_type='image/png'),
        ).status_code, 200)
        self.assertEqual(self.action('start_test').status_code, 200)
        self.assertEqual(self.action(
            'submit_test', test_raw_data='振动数据记录完整', test_conclusion_temp='试验合格',
        ).status_code, 200)
        self.assertEqual(self.action('issue_report', final_conclusion='检测项目全部合格').status_code, 200)

        action_codes = set(
            WorkflowEvent.objects.filter(order=self.order, actor=self.operator).exclude(action_code='')
            .values_list('action_code', flat=True)
        )
        self.assertTrue({
            'lab_schedule_assign', 'lab_test_start',
            'lab_test_submit', 'lab_report_issue',
        }.issubset(action_codes))
        schedule_event = WorkflowEvent.objects.get(order=self.order, action_code='lab_schedule_assign')
        self.assertEqual(schedule_event.schedule, self.schedule)
        self.assertIn('device', schedule_event.change_data)
        self.assertIn('sample_arrived', schedule_event.change_data)
        self.assertEqual(self.order.experiments.get().test_operator, self.operator)

    def test_operator_cannot_start_test_before_sample_arrives(self):
        self.assertEqual(self.action(
            'schedule_assign', device_id=self.device.id,
            plan_start_time='2026-10-01', plan_end_time='2026-10-03', sample_arrived='false',
        ).status_code, 200)
        response = self.action('start_test')
        self.assertEqual(response.status_code, 400)
        self.assertIn('样品尚未到达', response.json()['error'])

    def test_arrived_status_requires_sample_photo(self):
        response = self.action(
            'schedule_assign', device_id=self.device.id,
            plan_start_time='2026-10-01', plan_end_time='2026-10-03', sample_arrived='true',
        )
        self.assertEqual(response.status_code, 400)
        self.assertIn('必须上传至少一张样品照片', response.json()['error'])
        self.schedule.refresh_from_db()
        self.assertFalse(self.schedule.sample_arrived)

    def test_operator_can_query_and_export_only_own_laboratory(self):
        self.client.force_login(self.operator)
        response = self.client.get(reverse('laboratory_orders'), {'lab_type': 1, 'keyword': '操作员'})
        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.json()['total'], 1)
        self.assertEqual(response.json()['items'][0]['order_no'], self.order.order_no)

        denied = self.client.get(reverse('laboratory_orders'), {'lab_type': 2})
        self.assertEqual(denied.status_code, 403)
        exported = self.client.get(reverse('laboratory_orders_export'), {
            'lab_type': 1, 'schedule_ids': str(self.schedule.id),
        })
        self.assertEqual(exported.status_code, 200)
        self.assertIn('spreadsheetml', exported['Content-Type'])
        workbook = load_workbook(BytesIO(exported.content), read_only=True)
        rows = list(workbook.active.iter_rows(values_only=True))
        self.assertEqual(rows[1][1], self.order.order_no)

    def test_operator_cannot_manage_devices(self):
        self.client.force_login(self.operator)
        response = self.client.post(
            reverse('lab_devices'),
            data={'device_code': 'DENIED-001', 'device_name': '不应创建'},
            content_type='application/json',
        )
        self.assertEqual(response.status_code, 403)
